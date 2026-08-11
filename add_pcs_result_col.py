"""
Insert a RED/GREEN result column before the current column J in the PCS testing checklist.

Logic (applied as Excel conditional formatting, auto-updates when readings are filled):
  After insert:  new col J = result  |  command values still D-I  |  equipment readings → K-P
  GREEN: all readings entered AND every commanded value's reading is within ±2
  RED:   any reading entered but at least one is outside ±2 (green rule takes priority)
  No colour: no readings entered yet
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

FILE_PATH = (
    r"C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD"
    r"\HyESys Dept\3. Hardware (PCS.BATT)\PCS"
    r"\sinosoar - current v2.1\onsite testing"
    r"\PCS 测试清单 testing checklist - original.xlsx"
)

wb = openpyxl.load_workbook(FILE_PATH)
ws = wb.active

# ── Insert blank column at position 10 (letter J) ────────────────────────────
ws.insert_cols(10)

# Set narrow width for the result column
ws.column_dimensions["J"].width = 3

# ── Conditional formatting ────────────────────────────────────────────────────
# After insert:
#   Command values  : D(4) E(5) F(6) G(7) H(8) I(9)   (unchanged)
#   Equipment reads : K(11) L(12) M(13) N(14) O(15) P(16)  (old J-O shifted +1)
#   New result col  : J(10)

green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# GREEN: at least one reading entered AND every non-empty command is within ±2
#   IF(ISNUMBER(D14), reading K must exist and |D-K|<=2, else pass)  — per pair
green_formula = (
    "AND("
    "OR(ISNUMBER(K14),ISNUMBER(L14),ISNUMBER(M14),ISNUMBER(N14),ISNUMBER(O14),ISNUMBER(P14)),"
    "IF(ISNUMBER(D14),AND(ISNUMBER(K14),ABS(D14-K14)<=2),TRUE),"
    "IF(ISNUMBER(E14),AND(ISNUMBER(L14),ABS(E14-L14)<=2),TRUE),"
    "IF(ISNUMBER(F14),AND(ISNUMBER(M14),ABS(F14-M14)<=2),TRUE),"
    "IF(ISNUMBER(G14),AND(ISNUMBER(N14),ABS(G14-N14)<=2),TRUE),"
    "IF(ISNUMBER(H14),AND(ISNUMBER(O14),ABS(H14-O14)<=2),TRUE),"
    "IF(ISNUMBER(I14),AND(ISNUMBER(P14),ABS(I14-P14)<=2),TRUE)"
    ")"
)

# RED: any reading present (green rule fires first when all pass, overriding this)
red_formula = (
    "OR(ISNUMBER(K14),ISNUMBER(L14),ISNUMBER(M14),"
    "ISNUMBER(N14),ISNUMBER(O14),ISNUMBER(P14))"
)

cf_range = "J14:J51"

# Green rule added first → highest priority in Excel
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(formula=[green_formula], fill=green_fill, stopIfTrue=True),
)
ws.conditional_formatting.add(
    cf_range,
    FormulaRule(formula=[red_formula], fill=red_fill, stopIfTrue=True),
)

wb.save(FILE_PATH)
print(f"Saved: {FILE_PATH}")
