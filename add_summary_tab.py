import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Singapore DC Battery Cell Comparison.xlsx'
wb = openpyxl.load_workbook(path)

# Remove existing tab if present
if 'Cell Type Comparison' in wb.sheetnames:
    del wb['Cell Type Comparison']

ws = wb.create_sheet('Cell Type Comparison')

# ── Style helpers ────────────────────────────────────────────────────────────
def fill(hex_): return PatternFill('solid', fgColor=hex_)
def font(bold=False, size=11, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name='Calibri')
def align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin  = Side(style='thin',   color='BDBDBD')
med   = Side(style='medium', color='9E9E9E')
def border(top=None, bot=None, left=None, right=None):
    return Border(top=top or thin, bottom=bot or thin,
                  left=left or thin, right=right or thin)

BLUE_DARK  = '1565C0'
BLUE_MID   = '1976D2'
BLUE_LIGHT = 'E3F2FD'
GREEN_DARK = '2E7D32'
GREEN_MID  = '388E3C'
GREEN_LIGHT= 'E8F5E9'
AMBER      = 'F57F17'
AMBER_LIGHT= 'FFF8E1'
RED_DARK   = 'B71C1C'
RED_LIGHT  = 'FFEBEE'
GREY_HEAD  = '37474F'
GREY_LIGHT = 'F5F5F5'
WHITE      = 'FFFFFF'
WARN_YELL  = 'FFF176'

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {1: 4, 2: 32, 3: 80, 4: 4}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Helper: write a cell ──────────────────────────────────────────────────────
def w(row, col, value, bold=False, size=11, color='000000', bg=None,
      italic=False, h='left', v='center', wrap=True, border_=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if border_:
        c.border = border_
    return c

def merge_row(row, col_start, col_end, value, bold=False, size=11,
              color='000000', bg=None, h='center', italic=False):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.font      = font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = align(h=h, v='center', wrap=True)
    if bg:
        c.fill = fill(bg)
    return c

# ── ROW COUNTER ────────────────────────────────────────────────────────────────
r = 1

# ── TITLE ──────────────────────────────────────────────────────────────────────
ws.row_dimensions[r].height = 36
merge_row(r, 1, 4, 'Battery Cell Type Comparison: Cylindrical vs Prismatic',
          bold=True, size=16, color=WHITE, bg=GREY_HEAD, h='center')
r += 1

ws.row_dimensions[r].height = 22
merge_row(r, 1, 4,
          'Based on cell data in this workbook — verified against geometry and datasheet values',
          bold=False, size=10, color='757575', bg='ECEFF1', h='center',
          italic=True)
r += 1

# ── SECTION: CYLINDRICAL ──────────────────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 28
merge_row(r, 1, 4, 'CYLINDRICAL CELLS  (18650 · 21700 · 32140)',
          bold=True, size=13, color=WHITE, bg=BLUE_MID, h='left')
r += 1

# Sub-header: cells in dataset
ws.row_dimensions[r].height = 18
merge_row(r, 2, 4,
          'Cells in dataset: INR18650-2500A | INR21700-50E (×2) | INR21700-M50LT | NCR21700A | '
          'US21700VTC6A | Power Cell 32140 (Na-ion)',
          bold=False, size=9, color='455A64', bg=BLUE_LIGHT, h='left',
          italic=True)
r += 1

# Density note
ws.row_dimensions[r].height = 18
merge_row(r, 2, 4,
          'Cell-level kWh/m³ in dataset:  Li-ion 21700: 554–742  |  18650: 554  |  Na-ion 32140: 227  '
          '(lower energy chemistry)  |  Prismatic range for context: 345–440',
          bold=False, size=9, color=RED_DARK, bg=BLUE_LIGHT, h='left',
          italic=True)
r += 1

# Pros header
ws.row_dimensions[r].height = 20
merge_row(r, 2, 4, '  PROS', bold=True, size=11, color=WHITE, bg=GREEN_MID)
r += 1

cyl_pros = [
    ('Highest cell-level volumetric energy density',
     'Li-ion 21700 cells: 554–742 kWh/m³ (cell level) — highest of all cell types in this dataset. '
     'Even after 20–35% packing loss for cooling channels, pack-level density (390–550 kWh/m³) '
     'remains above prismatic packs.'),
    ('Mature high-volume manufacturing',
     'Widest supplier base; 18650 and 21700 form factors are global standards. Multi-sourcing '
     'reduces supply risk and keeps cost per Wh competitive.'),
    ('Strong mechanical casing',
     'Steel can withstands internal pressure from gas generation. Cylindrical geometry '
     'distributes hoop stress uniformly — robust under abuse and thermal runaway.'),
    ('Radial heat dissipation',
     'Heat generated at cell core radiates outward in all radial directions to the '
     'outer surface — efficient air or liquid cooling around each cell.'),
    ('Established safety data',
     'Extensive UL9540A, IEC 62133, and UN38.3 test data available. Failure modes '
     'well-characterised for 18650 and 21700 formats.'),
]

for title, detail in cyl_pros:
    ws.row_dimensions[r].height = 42
    w(r, 2, f'✔  {title}', bold=True, size=10, color=GREEN_DARK, bg=GREEN_LIGHT)
    w(r, 3, detail,        bold=False, size=10, color='212121',   bg=WHITE)
    r += 1

# Cons header
ws.row_dimensions[r].height = 20
merge_row(r, 2, 4, '  CONS', bold=True, size=11, color=WHITE, bg=RED_DARK)
r += 1

cyl_cons = [
    ('Packing inefficiency in modules',
     'Round cross-section leaves 20–35% of module volume as air gaps once cooling channels '
     'and structural spacers are included (geometric min 9–22%; practical with thermal '
     'management: 20–35%).'),
    ('High cell count per pack',
     'A 100 kWh pack using 21700 cells (~18 Wh each) requires ~5,500 cells. Each cell '
     'needs individual interconnects, nickel strips, and a BMS channel — assembly '
     'complexity and interconnect failure risk scale with cell count.'),
    ('BMS complexity',
     'Hundreds to thousands of cells require per-cell or per-group voltage monitoring. '
     'More channels, more firmware logic, more potential single-point failure nodes.'),
    ('Harder mid-pack fault detection',
     'A single failing cell buried inside a module is difficult to detect thermally '
     'or electrically without individual cell-level sensors.'),
]

for title, detail in cyl_cons:
    ws.row_dimensions[r].height = 42
    w(r, 2, f'✘  {title}', bold=True, size=10, color=RED_DARK,  bg=RED_LIGHT)
    w(r, 3, detail,        bold=False, size=10, color='212121',  bg=WHITE)
    r += 1

# ── SECTION: PRISMATIC ────────────────────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 28
merge_row(r, 1, 4, 'PRISMATIC CELLS',
          bold=True, size=13, color=WHITE, bg=AMBER, h='left')
r += 1

ws.row_dimensions[r].height = 18
merge_row(r, 2, 4,
          'Cells in dataset: NMC 50Ah (CATL) | 72174L4-280Ah (Batterotech) | '
          '72174L4-314Ah (Batterotech) | LF168 (CATL)',
          bold=False, size=9, color='455A64', bg=AMBER_LIGHT, h='left',
          italic=True)
r += 1

ws.row_dimensions[r].height = 18
merge_row(r, 2, 4,
          'Cell-level kWh/m³ in dataset: 345–440  (lower than cylindrical at cell level; '
          'advantage lies in pack-level simplicity and thermal stability, not raw density)',
          bold=False, size=9, color=RED_DARK, bg=AMBER_LIGHT, h='left',
          italic=True)
r += 1

# Pros header
ws.row_dimensions[r].height = 20
merge_row(r, 2, 4, '  PROS', bold=True, size=11, color=WHITE, bg=GREEN_MID)
r += 1

pri_pros = [
    ('High pack-level volumetric packing efficiency',
     'Flat surfaces stack with 85–92% volumetric efficiency — far less wasted space per '
     'module compared to cylindrical (65–80%). Reduces enclosure and rack footprint for '
     'a given pack capacity.'),
    ('Fewer cells per pack',
     'A 100 kWh pack using 280Ah LFP cells needs only ~35 cells (vs ~5,500 for 21700). '
     'Fewer interconnects, fewer solder/weld joints, and lower BMS channel count '
     'reduces assembly cost and fault probability.'),
    ('Simpler BMS architecture',
     'Low cell count → straightforward series-parallel topology. Each cell can be '
     'individually monitored with minimal hardware overhead.'),
    ('Industry-standard for stationary BESS',
     'Large-format prismatic LFP is the dominant format for grid-scale and data centre '
     'energy storage globally (CATL, BYD, EVE, Batterotech). Proven supply chain, '
     'module designs, and installation procedures.'),
    ('Easier field replacement',
     'Individual cells or modules are large and accessible — a failing cell in a BESS '
     'cabinet can be swapped without disassembling hundreds of interconnected cylindrical '
     'cell sub-packs.'),
    ('LFP thermal stability (large-format cells in dataset)',
     'LFP thermal runaway onset ~270°C vs ~150–180°C for NMC cylindrical cells. '
     'Reduced risk of thermal propagation in a multi-cell rack.'),
]

for title, detail in pri_pros:
    ws.row_dimensions[r].height = 42
    w(r, 2, f'✔  {title}', bold=True, size=10, color=GREEN_DARK, bg=GREEN_LIGHT)
    w(r, 3, detail,        bold=False, size=10, color='212121',   bg=WHITE)
    r += 1

# Cons header
ws.row_dimensions[r].height = 20
merge_row(r, 2, 4, '  CONS', bold=True, size=11, color=WHITE, bg=RED_DARK)
r += 1

pri_cons = [
    ('Lower cell-level volumetric energy density',
     'Prismatic cells in this dataset: 345–440 kWh/m³ vs 554–742 kWh/m³ for Li-ion '
     'cylindrical. Pack-level packing efficiency partially closes the gap but does not '
     'fully offset the cell-level disadvantage.'),
    ('Cell swelling under cycling',
     'Electrochemical expansion during charge/discharge causes prismatic cells to '
     'swell (LFP ~1–3% per cycle). Packs require mechanical compression fixtures, '
     'expansion gaps, and periodic re-torqueing of end plates.'),
    ('Thermal hotspots in large-format cells',
     'Heat generated at the cell core has a longer conduction path to the cooling surface '
     'in thick prismatic cells (e.g. 72mm thickness for 280Ah). Centre of cell runs '
     'hotter than edges — limits maximum discharge rate.'),
    ('Less robust casing under internal pressure',
     'Aluminium or polymer prismatic cases are less mechanically resistant to internal '
     'pressure spikes than cylindrical steel cans. Vent designs must be carefully '
     'engineered to prevent case rupture.'),
    ('Larger gas volume per cell in thermal runaway',
     'Large-format cells vent significantly more gas per event. Per UL9540A test data in '
     'this project: Batterotech 280Ah LFP cell vents 129.5 L/cell at 131.1°C surface '
     'temperature. A 1 MWh BESS (91 cells) releases ~11,785 L total gas — critical '
     'design input for ventilation and gas management systems.'),
    ('Reduced multi-sourcing flexibility',
     'Large-format prismatic form factors are less standardised than 18650/21700. '
     'Switching suppliers often requires re-validation of module mechanical design.'),
]

for title, detail in pri_cons:
    ws.row_dimensions[r].height = 48
    w(r, 2, f'✘  {title}', bold=True, size=10, color=RED_DARK,  bg=RED_LIGHT)
    w(r, 3, detail,        bold=False, size=10, color='212121',  bg=WHITE)
    r += 1

# ── ENERGY DENSITY NOTE ────────────────────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 28
merge_row(r, 1, 4, 'ENERGY DENSITY — KEY CLARIFICATION',
          bold=True, size=12, color=WHITE, bg=GREY_HEAD)
r += 1

note_lines = [
    ('Cell-level kWh/m³',
     'Cylindrical 21700 (Li-ion): 554–742 kWh/m³     Prismatic (NMC/LFP): 345–440 kWh/m³\n'
     'Cylindrical cells have higher cell-level volumetric density. This is confirmed by '
     'geometry calculations from the measured dimensions in this workbook.'),
    ('Pack-level kWh/m³ (estimated)',
     'Cylindrical at 65–80% packing efficiency: ~360–590 kWh/m³\n'
     'Prismatic at 85–92% packing efficiency:   ~294–405 kWh/m³\n'
     'Cylindrical packs remain comparable or higher in volumetric density even after '
     'packing losses. Earlier claim that "prismatic is higher at pack level" was incorrect '
     'and has been removed from this summary.'),
    ('Why prismatic dominates BESS despite lower density',
     'System simplicity: fewer cells, fewer interconnects, simpler BMS, easier '
     'maintenance, and established BESS supply chain — not energy density.'),
]

for label, detail in note_lines:
    ws.row_dimensions[r].height = 60
    w(r, 2, label,  bold=True,  size=10, color='212121', bg=WARN_YELL)
    w(r, 3, detail, bold=False, size=10, color='212121', bg='FFFFF0')
    r += 1

# ── FOOTER ─────────────────────────────────────────────────────────────────────
r += 1
ws.row_dimensions[r].height = 16
merge_row(r, 1, 4,
          'Source: Cell specifications from Cell Comparison tab | Geometry calculations: this workbook | '
          'Thermal runaway data: UL9540A Report (Batterotech 280Ah) | Prepared by HyESys Agent',
          bold=False, size=8, color='9E9E9E', bg='FAFAFA', h='center', italic=True)

wb.save(path)
print('Tab created and saved.')
