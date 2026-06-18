"""
Add UPS Photos tab to Singapore DC UPS Comparison.xlsx
Downloads product images for each UPS brand and inserts them neatly.
"""

import os
import io
import requests
from PIL import Image as PILImage
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = (
    r"C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\HyESys Dept"
    r"\3. Hardware (PCS.BATT)\v2.2 - data center\UPS"
    r"\Singapore DC UPS Comparison.xlsx"
)

IMG_DIR = r"C:\Users\JasonOng\AST_Agent\ups_images"
os.makedirs(IMG_DIR, exist_ok=True)

# ─── UPS image sources ───────────────────────────────────────────────────────
UPS_IMAGES = [
    {
        "brand":   "Schneider Electric",
        "series":  "Galaxy VS",
        "model":   "GVSUPS120KHS",
        "url":     "https://nationwidepower.com/wp-content/uploads/2024/12/GalaxyVS_LargeInt_Left-gallery.png",
        "filename": "schneider_galaxy_vs.png",
    },
    {
        "brand":   "Eaton",
        "series":  "93PM",
        "model":   "93PM 100/200 kW",
        "url":     "https://cdn.blueally.com/eatonguard/images/93pm/93pm_front_angled.jpg",
        "filename": "eaton_93pm.jpg",
    },
    {
        "brand":   "Huawei",
        "series":  "UPS5000-E",
        "model":   "UPS5000-E-120K-FM",
        "url":     "https://digitalpower.huawei.com/attachments/data_center_energy_new/72bcec5780e1426596d4b30e093433fc.jpeg",
        "filename": "huawei_ups5000e.jpg",
    },
    {
        "brand":   "Delta Electronics",
        "series":  "Ultron HPH",
        "model":   "HPH-120K",
        "url":     "https://www.deltapowersolutions.com/media/download/Ultron-HPH-series-UPS-60-120kW.jpg.jpg",
        "filename": "delta_ultron_hph.jpg",
    },
    {
        "brand":   "Vertiv",
        "series":  "Liebert EXL S1",
        "model":   "EXL S1 250 kW",
        "url":     "https://www.vertiv.com/49bd88/globalassets/products/critical-power/uninterruptible-power-supplies-ups/liebert-exl-s1-ups-300-1250-kw/cp-ups-emea-508x635-exl-lsa-family-hero.png",
        "filename": "vertiv_liebert_exl_s1.png",
    },
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# ─── Download + normalise images ─────────────────────────────────────────────
def download_image(entry):
    """Download image, convert to PNG, return local path. Returns None on failure."""
    out_path = os.path.join(IMG_DIR, entry["filename"])
    png_path = out_path.rsplit(".", 1)[0] + ".png"

    if os.path.exists(png_path):
        print(f"  [cached] {entry['brand']}")
        return png_path

    print(f"  Downloading: {entry['brand']} — {entry['url'][:80]}")
    try:
        resp = requests.get(entry["url"], headers=HEADERS, timeout=20)
        resp.raise_for_status()
        img = PILImage.open(io.BytesIO(resp.content)).convert("RGBA")
        # Resize to max 400px wide, maintaining aspect ratio
        max_w = 400
        if img.width > max_w:
            ratio = max_w / img.width
            img = img.resize((max_w, int(img.height * ratio)), PILImage.LANCZOS)
        # Convert RGBA → RGB with white background for Excel compatibility
        bg = PILImage.new("RGB", img.size, (255, 255, 255))
        bg.paste(img, mask=img.split()[3] if img.mode == "RGBA" else None)
        bg.save(png_path, "PNG")
        print(f"    Saved: {png_path}")
        return png_path
    except Exception as e:
        print(f"    FAILED ({e})")
        return None


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    s = Side(border_style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)

def make_thick_border():
    s = Side(border_style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── Add photos sheet ─────────────────────────────────────────────────────────
def add_photos_sheet(wb, downloaded):
    # Remove existing sheet if present
    if "UPS Photos" in wb.sheetnames:
        del wb["UPS Photos"]

    ws = wb.create_sheet("UPS Photos")

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title = ws["A1"]
    title.value = "UPS PRODUCT PHOTOS — Singapore Data Centre Models (Comparison Reference)"
    title.fill      = make_fill("0D2137")
    title.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = (
        "Images sourced from official manufacturer pages and authorised distributors  |  "
        "For reference only — actual unit appearance may vary by configuration  |  "
        "Compiled by HyESys Agent, June 2026"
    )
    sub.fill      = make_fill("1F3864")
    sub.font      = Font(name="Calibri", size=9, italic=True, color="CCDDEE")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Layout: 2 images per row, with spacer column between
    # Columns: A(spacer) | B-D(img1) | E(spacer) | F-H(img2) | I(spacer)
    # Set column widths
    COL_WIDTHS = {
        "A": 2,    # left margin
        "B": 5,    # img label col
        "C": 40,   # image col (wide)
        "D": 5,    # right of img1
        "E": 4,    # middle spacer
        "F": 5,    # img label col 2
        "G": 40,   # image col 2
        "H": 5,    # right of img2
        "I": 2,    # right margin
    }
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Each UPS block: label rows + image rows
    # Row pattern per UPS pair (starting at row R):
    #   R+0: brand header (merged C or G)
    #   R+1..R+20: image rows (20 rows ≈ 150px height @7.5pt/row)
    #   R+21: model / series label
    #   R+22: key spec summary
    #   R+23: blank spacer

    BRAND_COLORS = [
        ("0070C0", "FFFFFF"),  # Schneider — blue
        ("C00000", "FFFFFF"),  # Eaton — red
        ("CC0000", "FFFFFF"),  # Huawei — red
        ("1F7A3B", "FFFFFF"),  # Delta — green
        ("7030A0", "FFFFFF"),  # Vertiv — purple
    ]

    KEY_SPECS = [
        "100–150 kW  |  480V DC  |  96% eff  |  VRLA / Li-ion",
        "20–400 kW modular  |  480V DC  |  97% eff  |  VRLA / Li-ion",
        "30–800 kVA modular  |  320–576V DC  |  97% eff  |  VRLA / LFP",
        "60–120 kW  |  ±240V DC (480V)  |  96% eff  |  VRLA",
        "250–1200 kW  |  540V DC  |  96% eff  |  VRLA / LFP",
    ]

    IMG_ROW_HEIGHT  = 8    # points per row in image zone
    IMG_ROWS        = 22   # number of rows for image zone ≈ 176px
    LABEL_ROW_H     = 20
    SPEC_ROW_H      = 18
    SPACER_ROW_H    = 14

    # Image columns for left and right panels
    IMG_COLS = [("B", "C"), ("F", "G")]   # (label col, image col) per position

    current_row = 3  # start after title rows

    pairs = [
        (downloaded[0], downloaded[1]),
        (downloaded[2], downloaded[3]),
        (downloaded[4], None),
    ]

    for pair_idx, (left, right) in enumerate(pairs):
        items = [(left, 0), (right, 1)] if right else [(left, 0)]

        # Determine row span
        total_rows = 2 + IMG_ROWS + 2 + 1  # header + img + labels + spacer = 27 rows

        # Pre-set row heights for this block
        ws.row_dimensions[current_row].height     = LABEL_ROW_H
        for r in range(current_row + 1, current_row + 1 + IMG_ROWS):
            ws.row_dimensions[r].height = IMG_ROW_HEIGHT
        ws.row_dimensions[current_row + 1 + IMG_ROWS].height     = LABEL_ROW_H
        ws.row_dimensions[current_row + 2 + IMG_ROWS].height     = SPEC_ROW_H
        ws.row_dimensions[current_row + 3 + IMG_ROWS].height     = SPACER_ROW_H

        for item, side_idx in items:
            if item is None:
                continue

            label_col, img_col = IMG_COLS[side_idx]
            label_col_idx = ord(label_col) - ord("A") + 1
            img_col_idx   = ord(img_col)   - ord("A") + 1
            brand_idx = pair_idx * 2 + side_idx
            bg_hex, fg_hex = BRAND_COLORS[brand_idx]

            # Brand header row
            ws.merge_cells(
                start_row=current_row, start_column=label_col_idx,
                end_row=current_row, end_column=img_col_idx
            )
            hdr = ws.cell(row=current_row, column=label_col_idx,
                          value=f"{item['brand']}  |  {item['series']}")
            hdr.fill      = make_fill(bg_hex)
            hdr.font      = Font(name="Calibri", size=12, bold=True, color=fg_hex)
            hdr.alignment = Alignment(horizontal="center", vertical="center")
            hdr.border    = make_thick_border()

            # Image zone background
            for r in range(current_row + 1, current_row + 1 + IMG_ROWS):
                for c in [label_col_idx, img_col_idx]:
                    cell = ws.cell(row=r, column=c)
                    cell.fill   = make_fill("F7FBFF")
                    cell.border = make_border()

            # Insert image
            if item.get("local_path") and os.path.exists(item["local_path"]):
                xl_img = XLImage(item["local_path"])
                # Target: fit in img_col, IMG_ROWS rows
                # Each row ~8pt ≈ 10.7px; total height = IMG_ROWS * 10.7
                target_h_px = IMG_ROWS * 8 * 1.333   # pt → px approx
                target_w_px = 40 * 7.5               # col width 40 × 7.5px/unit
                # Scale to fit within box maintaining aspect ratio
                pil = PILImage.open(item["local_path"])
                orig_w, orig_h = pil.size
                scale = min(target_w_px / orig_w, target_h_px / orig_h)
                xl_img.width  = int(orig_w * scale)
                xl_img.height = int(orig_h * scale)
                # Anchor: start at image col, row after header
                from openpyxl.utils.cell import coordinate_from_string
                anchor_cell = f"{img_col}{current_row + 1}"
                xl_img.anchor = anchor_cell
                ws.add_image(xl_img)
            else:
                # Placeholder text if image failed
                mid_r = current_row + 1 + IMG_ROWS // 2
                ws.merge_cells(
                    start_row=mid_r, start_column=label_col_idx,
                    end_row=mid_r, end_column=img_col_idx
                )
                ph = ws.cell(row=mid_r, column=label_col_idx,
                             value="[Image unavailable — see manufacturer website]")
                ph.font      = Font(name="Calibri", size=9, italic=True, color="888888")
                ph.alignment = Alignment(horizontal="center", vertical="center")

            # Model label
            model_row = current_row + 1 + IMG_ROWS
            ws.merge_cells(
                start_row=model_row, start_column=label_col_idx,
                end_row=model_row, end_column=img_col_idx
            )
            model_cell = ws.cell(row=model_row, column=label_col_idx,
                                 value=f"Model: {item['model']}")
            model_cell.fill      = make_fill(bg_hex)
            model_cell.font      = Font(name="Calibri", size=10, bold=True, color=fg_hex)
            model_cell.alignment = Alignment(horizontal="center", vertical="center")
            model_cell.border    = make_thick_border()

            # Spec summary
            spec_row = current_row + 2 + IMG_ROWS
            ws.merge_cells(
                start_row=spec_row, start_column=label_col_idx,
                end_row=spec_row, end_column=img_col_idx
            )
            spec_cell = ws.cell(row=spec_row, column=label_col_idx,
                                value=KEY_SPECS[brand_idx])
            spec_cell.fill      = make_fill("EBF3FB")
            spec_cell.font      = Font(name="Calibri", size=8, color="1F3864")
            spec_cell.alignment = Alignment(horizontal="center", vertical="center",
                                            wrap_text=True)
            spec_cell.border    = make_border()

        current_row += 3 + IMG_ROWS  # advance to next pair

    # Footer note
    ws.merge_cells(f"A{current_row}:I{current_row}")
    footer = ws.cell(row=current_row, column=1,
        value=(
            "Image sources: nationwidepower.com (Schneider)  |  alpinepowersystems.com (Eaton, Vertiv)  |  "
            "digitalpower.huawei.com (Huawei)  |  deltapowersolutions.com (Delta)  |  "
            "For reference only — © respective manufacturers"
        ))
    footer.fill      = make_fill("1F3864")
    footer.font      = Font(name="Calibri", size=8, italic=True, color="AABBCC")
    footer.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[current_row].height = 16


def main():
    print("Downloading UPS product images...")
    downloaded = []
    for entry in UPS_IMAGES:
        local_path = download_image(entry)
        downloaded.append({**entry, "local_path": local_path})

    print("\nLoading Excel workbook...")
    wb = load_workbook(EXCEL_PATH)

    print("Adding UPS Photos sheet...")
    add_photos_sheet(wb, downloaded)

    wb.save(EXCEL_PATH)
    print(f"\nDone. Saved: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
