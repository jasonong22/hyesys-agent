import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Phase Change Coolant Comparison.xlsx'
wb = openpyxl.Workbook()

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, size=10, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def aln(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin  = Side(style='thin',   color='B0BEC5')
med   = Side(style='medium', color='78909C')
bdr   = Border(top=thin, bottom=thin, left=thin, right=thin)
bdr_m = Border(top=med,  bottom=med,  left=med,  right=med)

def c(ws, row, col, val='', bold=False, size=10, fg='000000', bg=None,
      italic=False, h='center', v='center', wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = fnt(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = aln(h=h, v=v, wrap=wrap)
    if bg: cell.fill = fill(bg)
    if border: cell.border = bdr
    return cell

def merge(ws, r, c1, c2, val='', bold=False, size=10, fg='000000', bg=None,
          h='center', italic=False, v='center'):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=val)
    cell.font      = fnt(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = aln(h=h, v=v, wrap=True)
    if bg: cell.fill = fill(bg)
    cell.border = bdr_m
    return cell

def rh(ws, row, h): ws.row_dimensions[row].height = h

# ── Palette ───────────────────────────────────────────────────────────────────
TITLE_BG   = '0D47A1'
DIELIC_HDR = '1565C0'
DIELIC_ROW = 'E3F2FD'
DIELIC_ALT = 'EEF7FF'
PCM_HDR    = '2E7D32'
PCM_ROW    = 'E8F5E9'
PCM_ALT    = 'F1FBF1'
COL_HDR_BG = '37474F'
WARN_BG    = 'FFF3CD'
FOOT_BG    = 'F5F5F5'
WHITE      = 'FFFFFF'
GREY_LT    = 'ECEFF1'
NOTE_FG    = '546E7A'

# ════════════════════════════════════════════════════════════════════════════
# TAB 1: Comparison Table
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Coolant Comparison'

# Column layout: A=spacer | B=Vendor | C=Product | D=Class | E=Phase Type |
# F=Working Principle | G=Boil/Melt Pt | H=Latent Heat | I=TC liq | J=TC solid |
# K=Density liq | L=Density solid | M=Sp Heat | N=Dielec Str | O=GWP | P=Battery Notes
col_widths = {1:3, 2:18, 3:20, 4:18, 5:20, 6:48, 7:14, 8:14,
              9:14, 10:13, 11:14, 12:14, 13:14, 14:14, 15:12, 16:45, 17:3}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

NCOL = 16  # last data column

# ── Title ─────────────────────────────────────────────────────────────────────
r = 1
rh(ws, r, 36)
merge(ws, r, 1, 17, 'Phase Change Coolants for Battery Thermal Management — Brand & Specification Comparison',
      bold=True, size=15, fg=WHITE, bg=TITLE_BG)
r += 1
rh(ws, r, 18)
merge(ws, r, 1, 17,
      'Two-Phase Immersion Dielectric Fluids  |  Phase Change Materials (PCMs)  |  '
      'Sources: vendor datasheets, ResearchGate, PMC — compiled by HyESys Agent  |  June 2026',
      bold=False, size=9, fg=NOTE_FG, bg=GREY_LT, italic=True)
r += 1

# ── Column headers ────────────────────────────────────────────────────────────
headers = ['Vendor', 'Product / Grade', 'Chemical Class', 'Cooling Mechanism',
           'Working Principle', 'Boiling / Melting Pt (°C)', 'Latent Heat (kJ/kg)',
           'Thermal Cond. Liquid\n(W/m·K)', 'Thermal Cond. Solid\n(W/m·K)',
           'Density Liquid\n(g/cm³)', 'Density Solid\n(g/cm³)',
           'Sp. Heat Liquid\n(kJ/kg·K)', 'Dielectric\nStrength (kV)',
           'GWP\n(100-yr)', 'Battery Application & Notes']
r += 1
rh(ws, r, 44)
for i, hdr in enumerate(headers):
    c(ws, r, i+2, hdr, bold=True, size=9, fg=WHITE, bg=COL_HDR_BG, h='center', v='center')
ws.freeze_panes = f'B{r+1}'
r += 1

# ── Helper for data rows ───────────────────────────────────────────────────────
def data_row(ws, row, vendor, product, chem_class, mech, principle,
             bp_mp, latent, tc_liq, tc_solid, dens_liq, dens_solid,
             sp_heat, dielec, gwp, notes, bg1, bg2, alt=False):
    bg = bg2 if alt else bg1
    rh(ws, row, 50)
    vals = [vendor, product, chem_class, mech, principle, bp_mp, latent,
            tc_liq, tc_solid, dens_liq, dens_solid, sp_heat, dielec, gwp, notes]
    haligns = ['left','left','left','left','left',
               'center','center','center','center',
               'center','center','center','center','center','left']
    for i, (val, ha) in enumerate(zip(vals, haligns)):
        c(ws, row, i+2, val, bold=False, size=9, bg=bg, h=ha, v='top')

# ════════════════════════════════════════════════════════════════════════════
# SECTION 1: TWO-PHASE IMMERSION DIELECTRIC FLUIDS
# ════════════════════════════════════════════════════════════════════════════
rh(ws, r, 26)
merge(ws, r, 1, 17,
      '  SECTION 1  —  TWO-PHASE IMMERSION DIELECTRIC FLUIDS  '
      '(Liquid boils at battery surface; latent heat absorbs heat flux; vapour condenses at remote cooler)',
      bold=True, size=11, fg=WHITE, bg=DIELIC_HDR, h='left')
r += 1

PRINCIPAL_2P = 'Dielectric fluid fills enclosure; liquid contacts battery surface directly; boils on hot surface absorbing latent heat; vapour rises to chilled condenser, condenses, and recirculates by gravity or pump. No pump required in passive two-phase loop.'
PRINCIPAL_1P = 'Single-phase dielectric hydrocarbon oil; heat removed entirely by convection/conduction — no boiling. Pumped recirculation required. Higher thermal conductivity than fluorinated fluids but no latent heat mechanism.'

dielectric_data = [
    # vendor, product, class, mech, principle, bp, latent, tc_liq, tc_sol, d_liq, d_sol, cp, dielec, gwp, notes
    ('3M / Solventum', 'Novec 7000', 'HFE\n(Hydrofluoroether)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '34', '142', '0.075', 'N/A', '1.40', 'N/A', '~1.30', '>40', '370–420',
     'Lowest boiling Novec — best match for 20–40°C battery window. Being discontinued by Solventum. Replace with Opteon SF33.'),
    ('3M / Solventum', 'Novec 7100', 'HFE', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '61', '112', '0.069', 'N/A', '1.51', 'N/A', '1.18', '>40', '297–320',
     'Most widely deployed Novec grade for Li-ion battery and data-centre immersion cooling. 61°C well-matched to NMC/LFP surface temps.'),
    ('3M / Solventum', 'Novec 7200', 'HFE', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '76', '119', '0.068', 'N/A', '1.42', 'N/A', '—', '>40', '59',
     'Best GWP in the Novec two-phase line (GWP 59). Boiling point suits battery systems reaching 60–75°C. Preferred for new designs over 7100.'),
    ('3M / Solventum', 'Novec 7300', 'HFE', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '98', '102', '0.063', 'N/A', '1.66', 'N/A', '—', '>40', '210',
     'Higher boiling point — suited to power electronics or high-temperature applications. Less common for direct Li-ion battery cooling.'),
    ('3M / Solventum', 'Novec 649\n(FK-5-1-12)', 'Fluoroketone (FK)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '49', '88', '0.059', 'N/A', '1.60', 'N/A', '~1.10', '>40', '~1',
     'Ultra-low GWP (~1) — most environmentally acceptable fluorinated fluid. Fire suppression dual function. Drop-in replacement for Novec 7100. Being phased out by Solventum; Chemours Opteon 2P50 is successor.'),
    ('3M / Solventum', 'Fluorinert FC-72', 'PFC\n(Perfluorocarbon)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '56', '88', '0.057', 'N/A', '1.68', 'N/A', '—', '>40', '~9,300',
     'Legacy semiconductor cooling fluid. GWP 9,300 — NOT recommended for new battery system designs. Retained for reference only.'),
    ('3M / Solventum', 'Fluorinert FC-3283', 'PFC\n(Perfluorocarbon)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '128', '78', '0.066', 'N/A', '1.82', 'N/A', '—', '>40', '>10,000',
     'High-boiling PFC for extreme-temperature electronics. GWP >10,000. Not suitable for modern battery BESS applications — environmental regulations prohibit new use.'),
    ('Syensqo\n(Solvay)', 'Galden EV55', 'PFPE\n(Perfluoropolyether)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '55', '—', '0.065', 'N/A', '1.65', 'N/A', '—', '40', '~10,000*',
     'Low-boiling PFPE grade designed specifically for EV battery direct immersion cooling. 55°C closely matches Li-ion thermal window. Chemically inert to all battery materials. *GWP not officially published by Syensqo.'),
    ('Syensqo\n(Solvay)', 'Galden EV110', 'PFPE', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '110', '—', '0.065', 'N/A', '1.71', 'N/A', '—', '40', '~10,000*',
     'Higher boiling PFPE — designed for thermal runaway containment scenarios where fluid must tolerate cells reaching 100+°C before boiling. Dual use: normal cooling + safety buffer.'),
    ('Syensqo\n(Solvay)', 'Galden HT135', 'PFPE', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '135', '~69', '0.065', 'N/A', '1.72', 'N/A', '~0.96', '40', '~10,000*',
     'Original high-temperature PFPE grade; used for power electronics in BMS inverters and DC-DC converters. Not suited for direct Li-ion cell cooling due to high boiling point.'),
    ('Chemours', 'Opteon SF33\n(HFO-1336mzz-Z)', 'HFO\n(Hydrofluoroolefin)', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '33.4', '—', '—', 'N/A', '—', 'N/A', '—', 'None', '2',
     'Ultra-low GWP (2) direct replacement for Novec 7000. HFO chemistry; zero ODP; non-flammable. Recommended new-design alternative as Novec 7000 is discontinued. Chemours expanding production capacity.'),
    ('Chemours', 'Opteon 2P50\n(HFO blend)', 'HFO blend', 'Two-Phase\nImmersion',
     PRINCIPAL_2P, '~49', '—', '0.073', 'N/A', '—', 'N/A', '1.09', 'None', '10',
     'Next-generation replacement for Novec 649, PFPEs, and PFCs. Boiling ~49°C matches Li-ion operating window. Samsung and server OEMs approved. Commercial availability 2025. GWP 10 vs 9,300+ for PFC equivalents.'),
    ('Engineered Fluids', 'ElectroCool EC-100', 'Synthetic\nHydrocarbon', 'Single-Phase\nImmersion',
     PRINCIPAL_1P, 'No boiling\n(Flash pt 190°C)', 'N/A\n(single phase)', '0.138\n@0°C; 0.136@40°C', 'N/A', '0.82', 'N/A', '2.04@0°C\n2.21@40°C', '≥40', '0',
     'Zero GWP; 94%+ biodegradable; synthetic hydrocarbon. 2× higher thermal conductivity than fluorinated fluids. No latent heat but simpler single-phase loop. Suitable for direct cell immersion without fluorine risk.'),
    ('Engineered Fluids', 'ElectroCool EC-110', 'Synthetic\nHydrocarbon', 'Single-Phase\nImmersion',
     PRINCIPAL_1P, 'No boiling\n(Flash pt 193°C)', 'N/A\n(single phase)', '0.138\n@0°C; 0.136@40°C', 'N/A', '0.82', 'N/A', '2.06@0°C\n2.21@40°C', '≥60', '0',
     'Higher dielectric strength (≥60kV) variant of EC-100. 95%+ biodegradable. Pour point −57°C. 25-year shelf life (sealed). Preferred for new BESS designs seeking zero-GWP single-phase immersion.'),
    ('Engineered Fluids', 'AmpCool AC-110', 'Synthetic\nHydrocarbon', 'Single-Phase\nImmersion',
     PRINCIPAL_1P, 'No boiling\n(Flash pt >180°C)', 'N/A\n(single phase)', '~0.138', 'N/A', '~0.80', 'N/A', '~2.2', '≥40', '0',
     'Formulated specifically for EV batteries, motors, and inverters. Compatible with all Li-ion battery chemistries. Density ~0.80 g/cm³ — lighter than water, reduces pack weight. Direct cell, module, and pack immersion.'),
]

for i, row_data in enumerate(dielectric_data):
    data_row(ws, r, *row_data, DIELIC_ROW, DIELIC_ALT, alt=(i%2==1))
    r += 1

# ════════════════════════════════════════════════════════════════════════════
# SECTION 2: PHASE CHANGE MATERIALS (PCMs)
# ════════════════════════════════════════════════════════════════════════════
r += 1
rh(ws, r, 26)
merge(ws, r, 1, 17,
      '  SECTION 2  —  PHASE CHANGE MATERIALS (PCMs)  '
      '(Solid melts at target temperature; latent heat absorption maintains near-constant cell temperature)',
      bold=True, size=11, fg=WHITE, bg=PCM_HDR, h='left')
r += 1

PRINCIPAL_PCM = 'Solid PCM surrounds battery cells or is integrated in module housing. When cell surface reaches melting point, PCM melts absorbing latent heat — cell temperature is clamped near-constant during the phase transition. PCM resolidifies when heat source is removed, ready for next cycle. Passive — no pump, no power.'
PRINCIPAL_ENCAP = 'Paraffin core microencapsulated in polymer shell (10–15 wt%); capsules dispersed in water or dielectric fluid as pumpable slurry. Combines latent heat absorption of PCM with convective heat transfer of liquid cooling. Shell prevents liquid paraffin contamination of cooling circuit.'

pcm_data = [
    ('Rubitherm\n(Germany)', 'RT28HC', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '27–29\n(melting range)', '250', '0.2\n(both phases)', '0.2', '~0.77\n(liquid)', '0.88\n(solid)', '~2.0', 'N/A', 'N/A',
     'Highest latent heat in the RT line (250 J/g). Melting range exactly brackets 25–35°C ideal Li-ion window. Most cited PCM in battery BTM academic literature for ambient-temperature charging scenarios.'),
    ('Rubitherm\n(Germany)', 'RT35HC', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '34–36\n(melting range)', '240', '0.2\n(both phases)', '0.2', '~0.77', '0.88', '~2.0', 'N/A', 'N/A',
     'Most widely validated Rubitherm grade in battery BTM research papers. 34–36°C activation temperature matches peak Li-ion cell surface temp during fast charge. No sub-cooling; congruent melting.'),
    ('Rubitherm\n(Germany)', 'RT44HC', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '41–45\n(melting range)', '250', '0.2\n(both phases)', '0.2', '~0.77', '0.80–0.88', '~2.0', 'N/A', 'N/A',
     'Equal latent heat to RT28HC (250 J/g). Higher activation point — suitable for high-rate discharge cells or LFP where 40–45°C is acceptable before PCM engages. Also used in power electronics BTM.'),
    ('Rubitherm\n(Germany)', 'RT55', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '51–57\n(melting range)', '170', '0.2\n(both phases)', '0.2', '~0.77', '0.88', '~2.0', 'N/A', 'N/A',
     'Lower latent heat than HC grades (170 J/g). Activates near thermal runaway early-warning zone (>50°C). Used as a safety buffer layer in BESS thermal management — second line of defence after primary BTM.'),
    ('Rubitherm\n(Germany)', 'RT60', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '53–68\n(wide range)', '~160–180', '0.2\n(both phases)', '0.2', '~0.77', '0.88', '~2.0', 'N/A', 'N/A',
     'Wide 15°C melting range — energy absorbed gradually across 53–68°C band. Suited to applications where temperature gradients within the pack are large. Less isothermal than HC grades.'),
    ('Rubitherm\n(Germany)', 'RT64HC', 'Paraffin wax\n(Organic)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '60–65\n(melting range)', '~240–250\n(est. HC grade)', '0.2\n(both phases)', '0.2', '~0.77', '~0.88', '~2.0', 'N/A', 'N/A',
     'Newest HC grade (2025 datasheet). HC = High Capacity — ~25–30% more latent heat than non-HC at same temperature. Suitable for power electronics thermal runaway buffering applications.'),
    ('Croda\n(UK)', 'CrodaTherm 37', 'Bio-fatty acid\n(100% bio-based)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '36\n(melting pt)', '216', '~0.15–0.20\n(est.)', '~0.15–0.20', '0.812\n(liquid)', '—', '—', 'N/A', 'N/A',
     '100% renewable carbon (USDA Certified Biobased). Lower flammability than paraffin. 36°C activation matches Li-ion ideal operating window. Environmentally superior to petroleum paraffin. TC not published — estimated from fatty acid class.'),
    ('Croda\n(UK)', 'CrodaTherm 47', 'Bio-fatty acid\n(100% bio-based)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '47\n(melting pt)', '197', '~0.15–0.20\n(est.)', '~0.15–0.20', '0.829\n(liquid)', '—', '—', 'N/A', 'N/A',
     '100% renewable carbon. Lower latent heat (197 J/g) vs Rubitherm HC grades. Activates at 47°C — used for higher temperature tolerance applications or as a secondary thermal buffer layer.'),
    ('Croda\n(UK)', 'CrodaTherm 53', 'Bio-fatty acid\n(100% bio-based)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '52\n(melting pt)', '220', '~0.15–0.20\n(est.)', '~0.15–0.20', '0.829\n(liquid)', '—', '—', 'N/A', 'N/A',
     '100% bio-based; 220 J/g latent heat at 52°C — competitive with paraffin HC grades. Croda microencapsulated (ME) versions available for slurry coolant integration.'),
    ('Croda\n(UK)', 'CrodaTherm 60', 'Bio-fatty acid\n(100% bio-based)', 'Passive PCM\n(Solid↔Liquid)',
     PRINCIPAL_PCM, '60\n(melting pt)', '215', '~0.15–0.20\n(est.)', '~0.15–0.20', '0.824\n(liquid)', '—', '—', 'N/A', 'N/A',
     '100% bio-based at 60°C — near thermal runaway onset for LFP. Used as high-temperature passive safety layer in BESS enclosures. ME grades available for slurry applications.'),
    ('Microtek Labs\n(USA)', 'MPCM37-D', 'Microencapsulated\nParaffin (polymer shell)', 'Slurry PCM\n(Recirculating)',
     PRINCIPAL_ENCAP, '37\n(peak); onset 25.5°C', '110\n(encapsulated)', 'N/A\n(slurry)', 'N/A', '~0.9\n(particle)', 'N/A', '2.70–3.21\n(slurry)', 'N/A', 'N/A',
     'Enables pumpable PCM slurry — combines convective cooling with latent heat. 10–15wt% polymer shell reduces effective latent heat to 110 J/g (vs ~220 J/g bulk paraffin) but prevents paraffin leakage into coolant circuit. Mean particle 17–20µm.'),
]

for i, row_data in enumerate(pcm_data):
    data_row(ws, r, *row_data, PCM_ROW, PCM_ALT, alt=(i%2==1))
    r += 1

# ── Notes row ─────────────────────────────────────────────────────────────────
r += 1
rh(ws, r, 20)
merge(ws, r, 1, 17,
      '  NOTES:  *Syensqo Galden PFPE GWP — not officially published; peer-reviewed literature estimates ~2,820–10,000 depending on grade.  '
      '†CrodaTherm thermal conductivity not disclosed by Croda; estimated 0.15–0.20 W/m·K from bio-fatty acid literature.  '
      '†Latent heat column: J/g = kJ/kg (same numerical value).  '
      'TC = Thermal Conductivity.  All data from manufacturer datasheets, ChemPoint, ResearchGate, PMC — June 2026.',
      bold=False, size=8, fg=NOTE_FG, bg=WARN_BG, italic=True, h='left')

print('Tab 1 written.')

# ════════════════════════════════════════════════════════════════════════════
# TAB 2: Product Images — placeholder (will be updated with real images)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Product Images')
ws2.sheet_view.showGridLines = False

# Column widths for image grid: 4 columns of products
img_col_widths = {1:3, 2:30, 3:3, 4:30, 5:3, 6:30, 7:3, 8:30, 9:3}
for col, w in img_col_widths.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

rh(ws2, 1, 36)
merge(ws2, 1, 1, 9, 'Phase Change Coolants — Product Images',
      bold=True, size=15, fg=WHITE, bg=TITLE_BG)
rh(ws2, 2, 18)
merge(ws2, 2, 1, 9,
      'Representative product photos — sourced from vendor websites  |  '
      'Images will be inserted in a subsequent step',
      bold=False, size=9, fg=NOTE_FG, bg=GREY_LT, italic=True)

print('Tab 2 placeholder written.')

wb.save(path)
print(f'File saved: {path}')
