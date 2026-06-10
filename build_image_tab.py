import sys, os, urllib.request, tempfile
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Phase Change Coolant Comparison.xlsx'
wb = load_workbook(path)

# ── Replace Tab 2 ─────────────────────────────────────────────────────────────
if 'Product Images' in wb.sheetnames:
    del wb['Product Images']
ws = wb.create_sheet('Product Images', 1)
ws.sheet_view.showGridLines = False

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, size=10, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def aln(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin = Side(style='thin', color='CFD8DC')
def bdr(): return Border(top=thin, bottom=thin, left=thin, right=thin)

def ce(row, col, val='', bold=False, size=10, fg='000000', bg=None,
       italic=False, h='center', v='center', wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap)
    if bg: c.fill = fill(bg)
    c.border = bdr()
    return c

def mg(row, c1, c2, val='', bold=False, size=10, fg='000000', bg=None,
       h='center', italic=False, v='center'):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap=True)
    if bg: c.fill = fill(bg)
    c.border = bdr()
    return c

def rh(row, h): ws.row_dimensions[row].height = h
def bg_row(row, bg, ncol=9):
    for col in range(1, ncol+1):
        ws.cell(row=row, column=col).fill = fill(bg)

# ── Column widths ──────────────────────────────────────────────────────────────
# A=spacer, B=product-left, C=divider, D=product-right, E=spacer
col_w = {1:3, 2:44, 3:3, 4:44, 5:3}
for col, w in col_w.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Image download utility ─────────────────────────────────────────────────────
TMPDIR = tempfile.mkdtemp(prefix='coolant_imgs_')
downloaded = {}

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

def dl(key, url):
    ext = '.jpg' if '.jpg' in url.lower() or 'jpeg' in url.lower() else '.png'
    fpath = os.path.join(TMPDIR, f'{key}{ext}')
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15) as r, open(fpath, 'wb') as f:
            f.write(r.read())
        if os.path.getsize(fpath) > 1000:
            downloaded[key] = fpath
            print(f'  OK  {key}: {os.path.getsize(fpath)//1024} kB')
        else:
            print(f'  SKIP {key}: file too small')
    except Exception as e:
        print(f'  FAIL {key}: {e}')

def insert_img(key, anchor_cell, w=290, h=180):
    if key not in downloaded:
        return
    try:
        img = XLImage(downloaded[key])
        img.width  = w
        img.height = h
        img.anchor = anchor_cell
        ws.add_image(img)
    except Exception as e:
        print(f'  IMG INSERT FAIL {key}: {e}')

# ── Download all images ────────────────────────────────────────────────────────
print('Downloading images...')
dl('novec_7100',      'https://tmcindustries.com/cdn/shop/files/240626-3M_Novec_7100-Glass-Gallon.jpg?v=1719433544')
dl('galden_ht55',     'https://www.yamabala.com/upload/catalog_list_pic/enL_catalog_25F12_K6BY7iZv3z.jpg')
dl('opteon_sf33',     'https://tmcindustries.com/cdn/shop/files/SF33.png?v=1723479929&width=750')
dl('ampcool',         'https://shop.engineeredfluids.com/cdn/shop/products/AC-110B_web.jpg?v=1675264551')
dl('electrocool',     'https://shop.engineeredfluids.com/cdn/shop/products/EC-110_web.jpg?v=1675264540')
dl('rubitherm_rt',    'https://www.rubitherm.eu/media/products/images/_detailImage/Rubitherm-RT-400x240.jpg')
dl('croda_brochure',  'https://www.crodaindustrialspecialties.com/mediaassets/images/industrial-chemicals/brochure-front-covers/screenshot-20240215-132834.png?w=500&la=en-GB')
dl('microtek_micro',  'https://microteklabs.com/assets/microcapsules-micrograph-DjIT-j5b.jpg')
dl('app_immersion',   'https://www.emobility-engineering.com/content/uploads/2024/02/EME-Carrar-min-1024x640.jpg')
dl('app_pcm_cells',   'https://www.emobility-engineering.com/content/uploads/2024/02/EME-paraffin-cells-min-1024x331.jpg')
dl('app_pcm_module',  'https://www.emobility-engineering.com/content/uploads/2024/02/EME-compsite-PCM-min-1024x576.jpg')

# ── Palette ───────────────────────────────────────────────────────────────────
TITLE_BG  = '0D47A1'
DIEL_BG   = '1565C0'
PCM_BG    = '2E7D32'
APP_BG    = '4A148C'
LABEL_BG  = 'ECEFF1'
IMG_BG    = 'F5F5F5'
SPEC_BG   = 'E8EAF6'
SPEC_PCM  = 'E8F5E9'
SPEC_APP  = 'F3E5F5'
WHITE     = 'FFFFFF'
NOTE_FG   = '546E7A'
SEP_BG    = 'CFD8DC'

# ── Layout builder ────────────────────────────────────────────────────────────
IMG_H     = 160   # row height for image rows (points ≈ 213px)
IMG_W     = 290   # image display width (px)
IMG_DISP  = 165   # image display height (px)

def section_header(r, text, bg, ncols=5):
    rh(r, 28)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = fnt(True, 12, WHITE)
    c.alignment = aln('left', 'center')
    c.fill      = fill(bg)
    return r+1

def product_block(r, col_b, col_d,
                  left_key, left_vendor, left_prod, left_spec,
                  right_key, right_vendor, right_prod, right_spec,
                  spec_bg):
    """Write one row-block for two products side by side.
    Returns the next available row after this block.
    """
    # Image row
    rh(r, IMG_H)
    ws.cell(row=r, column=col_b).fill = fill(IMG_BG)
    ws.cell(row=r, column=col_d).fill = fill(IMG_BG)
    if left_key:  insert_img(left_key,  get_column_letter(col_b)+str(r), IMG_W, IMG_DISP)
    if right_key: insert_img(right_key, get_column_letter(col_d)+str(r), IMG_W, IMG_DISP)
    r += 1

    # Vendor row
    rh(r, 18)
    ce(r, col_b, left_vendor,  bold=True,  size=9, fg='1A237E', bg=LABEL_BG)
    ce(r, col_d, right_vendor, bold=True,  size=9, fg='1A237E', bg=LABEL_BG)
    r += 1

    # Product name row
    rh(r, 18)
    ce(r, col_b, left_prod,  bold=True,  size=10, fg='000000', bg=WHITE)
    ce(r, col_d, right_prod, bold=True,  size=10, fg='000000', bg=WHITE)
    r += 1

    # Spec row
    rh(r, 32)
    ce(r, col_b, left_spec,  bold=False, size=9, fg='212121', bg=spec_bg, v='top')
    ce(r, col_d, right_spec, bold=False, size=9, fg='212121', bg=spec_bg, v='top')
    r += 1

    # Spacer
    rh(r, 8)
    bg_row(r, 'FAFAFA')
    r += 1

    return r

PLACEHOLDER = 'No product photo available\n— see vendor website'

r = 1

# ── Title ──────────────────────────────────────────────────────────────────────
rh(r, 36)
mg(r, 1, 5, 'Phase Change Coolants — Product Images & Application Photos',
   bold=True, size=15, fg=WHITE, bg=TITLE_BG)
r += 1
rh(r, 18)
mg(r, 1, 5,
   'Product images sourced from vendor & distributor websites  |  Application photos: eMobility Engineering (© all rights reserved)  |  Compiled by HyESys Agent  |  June 2026',
   bold=False, size=8, fg=NOTE_FG, bg='ECEFF1', italic=True)
r += 1
rh(r, 6); bg_row(r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════════════════════
# SECTION 1: DIELECTRIC FLUIDS
# ════════════════════════════════════════════════════════════════════════════════
r = section_header(r, '  SECTION 1 — TWO-PHASE IMMERSION DIELECTRIC FLUIDS  (Liquid boils on cell surface; condenses at remote cooler)', DIEL_BG)

r = product_block(r, 2, 4,
    'novec_7100',
    '3M / Solventum',
    'Novec 7100  (HFE-7100)',
    'Chemical Class: HFE (Hydrofluoroether)\nBoiling Pt: 61°C  |  Latent Heat: 112 kJ/kg\nThermal Cond. (liq): 0.069 W/m·K  |  Density: 1.51 g/cm³\nGWP: 297–320  |  Dielectric: >40 kV\nStatus: Being discontinued → replace with Opteon SF33',

    'galden_ht55',
    'Syensqo (Solvay)',
    'Galden EV55 / HT55  (PFPE)',
    'Chemical Class: PFPE (Perfluoropolyether)\nBoiling Pt: 55°C  |  Latent Heat: — (not disclosed)\nThermal Cond. (liq): 0.065 W/m·K  |  Density: 1.65 g/cm³\nGWP: ~10,000*  |  Dielectric: 40 kV\nNote: *GWP not officially published by Syensqo',
    SPEC_BG)

r = product_block(r, 2, 4,
    'opteon_sf33',
    'Chemours',
    'Opteon SF33  (HFO-1336mzz-Z)',
    'Chemical Class: HFO (Hydrofluoroolefin)\nBoiling Pt: 33°C  |  GWP: 2 (ultra-low)\nDielectric: None (non-fluorinated route)  |  ODP: 0\nDirect replacement for Novec 7000\nStatus: Commercial — Chemours expanding capacity',

    None,
    'Chemours',
    'Opteon 2P50  (HFO blend)',
    'Chemical Class: HFO blend\nBoiling Pt: ~49°C  |  GWP: 10\nThermal Cond. (liq): 0.073 W/m·K  |  Sp. Heat: 1.09 kJ/kg·K\nDirect replacement for Novec 649 / PFPE\nStatus: Early commercial (2025); Samsung approved',
    SPEC_BG)

r = product_block(r, 2, 4,
    'ampcool',
    'Engineered Fluids',
    'AmpCool AC-110',
    'Chemical Class: Synthetic Hydrocarbon (GWP = 0)\nSingle-phase immersion (no boiling — pumped loop)\nThermal Cond.: ~0.138 W/m·K  |  Density: ~0.80 g/cm³\nDielectric: ≥40 kV  |  Flash Pt: >180°C\nBiodegradable ≥95%  |  Designed for EV battery packs',

    'electrocool',
    'Engineered Fluids',
    'ElectroCool EC-110',
    'Chemical Class: Synthetic Hydrocarbon (GWP = 0)\nSingle-phase immersion (no boiling — pumped loop)\nThermal Cond.: 0.136 W/m·K  |  Density: 0.82 g/cm³\nDielectric: ≥60 kV  |  Flash Pt: 193°C\nPour Pt: −57°C  |  Biodegradable ≥95%  |  Shelf life 25 yr',
    SPEC_BG)

rh(r, 6); bg_row(r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════════════════════
# SECTION 2: PHASE CHANGE MATERIALS
# ════════════════════════════════════════════════════════════════════════════════
r = section_header(r, '  SECTION 2 — PHASE CHANGE MATERIALS (PCMs)  (Solid melts at target temperature; absorbs latent heat; passive — no pump)', PCM_BG)

r = product_block(r, 2, 4,
    'rubitherm_rt',
    'Rubitherm Technologies GmbH  (Germany)',
    'RT Series — RT28HC / RT35HC / RT44HC',
    'Chemical Class: Paraffin wax (petroleum-based)\nMelting Range: 27–45°C (grade-dependent)\nLatent Heat: 240–250 J/g  |  Thermal Cond.: 0.2 W/m·K\nDensity solid: 0.88 g/cm³  |  GWP: N/A\nMost cited PCM brand in battery BTM research literature',

    'croda_brochure',
    'Croda Industrial Specialties  (UK)',
    'CrodaTherm Series — CrodaTherm 37 / 47 / 53 / 60',
    'Chemical Class: Bio-fatty acid (100% renewable carbon)\nMelting Range: 36–60°C (grade-dependent)\nLatent Heat: 197–220 J/g  |  TC: ~0.15–0.20 W/m·K (est.)\nDensity liq: 0.81–0.83 g/cm³  |  GWP: N/A\nUSDA Certified Biobased; lower flammability than paraffin',
    SPEC_PCM)

r = product_block(r, 2, 4,
    'microtek_micro',
    'Microtek Laboratories  (USA)',
    'mPCM37-D  (Microencapsulated Paraffin)',
    'Chemical Class: Paraffin core in polymer shell (10–15 wt%)\nMelting Peak: 37°C  |  Latent Heat: 110 J/g (encapsulated)\nSp. Heat: 2.70–3.21 kJ/kg·K  |  Particle size: 17–20 µm\nForm: Pumpable slurry (20–40 wt% in water or fluid)\nEnables recirculated PCM cooling loop',

    None,
    '',
    '',
    '',
    SPEC_PCM)

rh(r, 6); bg_row(r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════════════════════
# SECTION 3: APPLICATION PHOTOS
# ════════════════════════════════════════════════════════════════════════════════
r = section_header(r,
    '  SECTION 3 — APPLICATION PHOTOS  (How these technologies are deployed in battery thermal management)',
    APP_BG)

r = product_block(r, 2, 4,
    'app_immersion',
    'Two-Phase Immersion Cooling',
    'Battery Pack in Dielectric Fluid',
    'Battery cells or modules are fully submerged in\ndielectric fluid inside a sealed enclosure.\nFluid boils on the cell surface, absorbing heat.\nVapour condenses at a chilled plate and\ndrains back — no pump required in passive loop.',

    'app_pcm_cells',
    'PCM Passive Thermal Buffer',
    'Phase Change Material Around Battery Cells',
    'Solid PCM (e.g. Rubitherm RT35HC) fills the\nspace between battery cells in a module housing.\nCell heat melts the PCM at ~35°C, absorbing\nheat and clamping cell temperature.\nResolidifies when pack cools between cycles.',
    SPEC_APP)

r = product_block(r, 2, 4,
    'app_pcm_module',
    'Enhanced PCM with Metal Fins',
    'Composite PCM — Thermal Conductivity Enhancement',
    'PCM thermal conductivity (~0.2 W/m·K for paraffin)\nis insufficient for high-rate discharge.\nSolution: embed metal foam, copper fins,\nor graphite matrix in PCM to boost\neffective conductivity to 1–5 W/m·K.',

    None, '', '',
    'Image not available — refer to vendor website.\nApplications: EV battery packs, grid BESS\ncabinets, data centre UPS systems,\nand aerospace battery thermal management.',
    SPEC_APP)

# ── Footer ─────────────────────────────────────────────────────────────────────
rh(r, 6); bg_row(r, 'FAFAFA'); r += 1
rh(r, 16)
mg(r, 1, 5,
   'Image credits: Engineered Fluids shop  |  Yamabala.com (Syensqo Galden)  |  TMC Industries (Novec, Opteon)  |  Rubitherm GmbH  |  '
   'Croda Industrial Specialties  |  Microtek Laboratories  |  eMobility Engineering (application photos)  |  All images © respective owners',
   bold=False, size=8, fg='9E9E9E', bg='FAFAFA', h='center', italic=True)

wb.save(path)
print(f'\nSaved. Downloaded {len(downloaded)}/{11} images.')
print(f'Images in: {TMPDIR}')
