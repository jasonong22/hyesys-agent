"""
UPS Comparison Excel Builder — Singapore Data Centre Focus
Target: ~120 kWh battery capacity integration
Sources: Schneider Electric Galaxy VS, Eaton 93PM, Huawei UPS5000-E, Delta Ultron HPH,
         Vertiv Liebert EXL S1 — official datasheets & spec pages (June 2026)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

OUTPUT_PATH = (
    r"C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\HyESys Dept"
    r"\3. Hardware (PCS.BATT)\v2.2 - data center\UPS"
    r"\Singapore DC UPS Comparison.xlsx"
)

# ─── Colour palette ─────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E75B6"   # mid blue
C_SUBHDR_FG   = "FFFFFF"
C_SECTION_BG  = "D6E4F0"   # light blue
C_SECTION_FG  = "1F3864"
C_HIGHLIGHT   = "FFF2CC"   # pale yellow — 120 kWh range rows
C_ALT_ROW     = "F2F7FB"   # very light blue
C_WHITE       = "FFFFFF"
C_WARN_BG     = "FCE4D6"   # light orange — caution cells
C_GOOD_BG     = "E2EFDA"   # light green — good cells
C_BORDER      = "BDD7EE"
C_TITLE_BG    = "0D2137"   # near-black blue

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(color=C_BORDER, thin=True):
    s = Side(border_style="thin" if thin else "medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_thick_border():
    thick = Side(border_style="medium", color="1F3864")
    thin  = Side(border_style="thin",   color=C_BORDER)
    return Border(left=thick, right=thick, top=thin, bottom=thin)


# ─── UPS Data ────────────────────────────────────────────────────────────────
# Each dict represents one UPS model entry.
# All electrical values are for the 100–125 kW range (closest to 120 kWh target).
# Battery specs assume external battery cabinet for ~120 kWh capacity.
#
# DC Discharge Current calculated as:  kW_rated × 1000 / DC_nom_V
# DC Charge Current from official specs where available.
# Battery kWh per cabinet: 40 × 12V VRLA × 250 Ah = 120 kWh (standard config)
#                           OR Li-ion equivalent stated per manufacturer.
#
# Sources:
#  [SE]    Schneider Electric Galaxy VS IEC Technical Specifications (990-91141L_EN)
#  [EA]    Eaton 93PM 30-200 kW Technical Specification (APAC Rev 2.653)
#  [HW]    Huawei UPS5000-E-(30-120k)-FM Datasheet + Support Portal
#  [DE]    Delta Ultron HPH Series Spec Page (deltapowersolutions.com, June 2026)
#  [VT]    Vertiv Liebert EXL S1 Data Sheet (500-600 kVA, 250-1200 kW)

UPS_DATA = [
    # ── Schneider Electric Galaxy VS 100 kW (IEC 400V) ──────────────────────
    {
        "brand":             "Schneider Electric",
        "series":            "Galaxy VS",
        "model":             "GVSUPS100KHS",
        "topology":          "Double Conversion (Online)",
        "power_kw":          100,
        "power_kva":         100,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "40–70 Hz",
        "ac_in_current_a":   152,           # @ 100 kW, 380V: 100000/(380×√3) ≈ 152 A
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  152,
        "dc_nom_v":          480,
        "dc_min_v":          384,           # End of Discharge (32 × 12V)
        "dc_max_v":          576,           # Max charge (48 × 12V)
        "dc_discharge_a":    208,           # 100000 W ÷ 480 V
        "dc_charge_a":       20,            # Standard charger; 40A optional
        "batt_type":         "VRLA (standard) / Li-ion Smart Modular (option)",
        "batt_kwh_cabinet":  120,           # 40 × 12V × 250Ah = 120 kWh (external cab)
        "batt_cab_v":        "480 V DC (40 blocks × 12V)",
        "eff_online_pct":    96,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1750 × 450 × 800",
        "ups_weight_kg":     225,
        "batt_cab_dims":     "1750 × 600 × 900 (per external cabinet)",
        "batt_cab_weight_kg": 1100,         # ~1100 kg for 40 × 250Ah VRLA blocks
        "comms":             "SNMP (NMC card), Modbus TCP, RS-232, dry contacts",
        "sg_presence":       "Very High — standard in SG DCs",
        "suitable_120kwh":   True,
        "notes":             "Most widely deployed UPS in Singapore data centres. "
                             "120 kW model also available (GVSUPS120KHS). "
                             "Smart modular Li-ion option significantly reduces weight. "
                             "Source: [SE] 990-91141L_EN",
    },
    # ── Schneider Electric Galaxy VS 120 kW (IEC 400V) ──────────────────────
    {
        "brand":             "Schneider Electric",
        "series":            "Galaxy VS",
        "model":             "GVSUPS120KHS",
        "topology":          "Double Conversion (Online)",
        "power_kw":          120,
        "power_kva":         120,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "40–70 Hz",
        "ac_in_current_a":   182,           # 120000/(380×√3) ≈ 182 A
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  182,
        "dc_nom_v":          480,
        "dc_min_v":          384,
        "dc_max_v":          576,
        "dc_discharge_a":    250,           # 120000 ÷ 480
        "dc_charge_a":       20,
        "batt_type":         "VRLA (standard) / Li-ion Smart Modular (option)",
        "batt_kwh_cabinet":  120,
        "batt_cab_v":        "480 V DC (40 blocks × 12V)",
        "eff_online_pct":    96,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1750 × 450 × 800",
        "ups_weight_kg":     225,
        "batt_cab_dims":     "1750 × 600 × 900 (per external cabinet)",
        "batt_cab_weight_kg": 1100,
        "comms":             "SNMP (NMC card), Modbus TCP, RS-232, dry contacts",
        "sg_presence":       "Very High — standard in SG DCs",
        "suitable_120kwh":   True,
        "notes":             "120 kW = directly 1 hr runtime on 120 kWh battery. "
                             "Key target model for this comparison. "
                             "Source: [SE] 990-91141L_EN",
    },
    # ── Eaton 93PM 100 kW (IEC 400V) ────────────────────────────────────────
    {
        "brand":             "Eaton",
        "series":            "93PM",
        "model":             "93PM 100 kW (400V IEC)",
        "topology":          "Double Conversion (Online)",
        "power_kw":          100,
        "power_kva":         100,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "45–65 Hz",
        "ac_in_current_a":   152,
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  152,
        "dc_nom_v":          480,           # 40 × 12V (standard config)
        "dc_min_v":          432,           # EOD: 40 × 10.8V = 432V
        "dc_max_v":          564,           # Max charge: 240 cells × 2.35V = 564V
        "dc_discharge_a":    208,           # 100000 ÷ 480
        "dc_charge_a":       33,            # Initial scalable; 50A fully configured
        "batt_type":         "VRLA / Wet Cell / Li-ion (optional)",
        "batt_kwh_cabinet":  120,           # 40 × 12V × 250Ah = 120 kWh
        "batt_cab_v":        "480 V DC (40 × 12V blocks) or 432V (36 × 12V)",
        "eff_online_pct":    97,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1750 × 480 × 810",
        "ups_weight_kg":     210,
        "batt_cab_dims":     "1800 × 600 × 900 (per external cabinet)",
        "batt_cab_weight_kg": 1050,
        "comms":             "SNMP (Network-MS card), Modbus RTU/TCP, REPO, dry contacts",
        "sg_presence":       "High — common in enterprise and colo DCs",
        "suitable_120kwh":   True,
        "notes":             "Scalable modular architecture: start with 30 kW, add 10 kW power "
                             "modules up to 200 kW in same frame. Battery float: 497–552V. "
                             "Li-ion option reduces battery footprint ~50%. "
                             "Source: [EA] Eaton 93PM APAC Tech Spec Rev 2.653",
    },
    # ── Eaton 93PM 200 kW — for reference (2 × 60 kWh cabinets = 120 kWh) ──
    {
        "brand":             "Eaton",
        "series":            "93PM",
        "model":             "93PM 200 kW (400V IEC)",
        "topology":          "Double Conversion (Online)",
        "power_kw":          200,
        "power_kva":         200,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "45–65 Hz",
        "ac_in_current_a":   304,
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  304,
        "dc_nom_v":          480,
        "dc_min_v":          432,
        "dc_max_v":          564,
        "dc_discharge_a":    417,           # 200000 ÷ 480
        "dc_charge_a":       50,            # Fully configured
        "batt_type":         "VRLA / Wet Cell / Li-ion (optional)",
        "batt_kwh_cabinet":  120,           # Same 120 kWh cabinet — but gives only 36 min runtime
        "batt_cab_v":        "480 V DC (40 × 12V blocks)",
        "eff_online_pct":    97,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1750 × 480 × 810",
        "ups_weight_kg":     230,
        "batt_cab_dims":     "1800 × 600 × 900 (per external cabinet)",
        "batt_cab_weight_kg": 1050,
        "comms":             "SNMP (Network-MS card), Modbus RTU/TCP, REPO, dry contacts",
        "sg_presence":       "High",
        "suitable_120kwh":   False,
        "notes":             "At 200 kW, a 120 kWh battery gives only ~36 min runtime. "
                             "Two 120 kWh cabinets (240 kWh) recommended for this rating. "
                             "Source: [EA] Eaton 93PM APAC Tech Spec Rev 2.653",
    },
    # ── Huawei UPS5000-E 120 kVA (modular) ─────────────────────────────────
    {
        "brand":             "Huawei",
        "series":            "UPS5000-E",
        "model":             "UPS5000-E-120K-FM",
        "topology":          "Double Conversion (Online) — Modular",
        "power_kw":          120,
        "power_kva":         120,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "45–65 Hz",
        "ac_in_current_a":   182,
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  182,
        "dc_nom_v":          480,
        "dc_min_v":          320,           # Wide range: supports VRLA & Li-ion
        "dc_max_v":          576,
        "dc_discharge_a":    250,           # 120000 ÷ 480
        "dc_charge_a":       40,            # Adjustable; VRLA typically 0.1C
        "batt_type":         "VRLA / Li-ion LFP (SmartLi ESM-48100B1)",
        "batt_kwh_cabinet":  120,
        "batt_cab_v":        "480 V DC (configurable 320–576V)",
        "eff_online_pct":    97,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "2000 × 600 × 800 (4-module frame)",
        "ups_weight_kg":     280,
        "batt_cab_dims":     "2000 × 600 × 800 (SmartLi rack, 10 modules = 48 kWh)",
        "batt_cab_weight_kg": 720,          # 48 kWh LFP rack; 3 racks for 144 kWh
        "comms":             "SNMP, Modbus TCP, WebUI, Huawei NetEco integration",
        "sg_presence":       "High — hyperscale DCs (AWS, Alibaba SG) and Tier 3 DCs",
        "suitable_120kwh":   True,
        "notes":             "Fully modular: 30 kVA power modules hot-swappable. "
                             "SmartLi Li-ion battery (LFP): 48V/100Ah per module = 4.8 kWh; "
                             "10 modules/rack = 48 kWh; 3 racks needed for 144 kWh. "
                             "VRLA option: 40 × 12V × 250Ah = 120 kWh in one standard cabinet. "
                             "Wide DC range (320–576V) future-proofs for Li-ion chemistry. "
                             "Source: [HW] UPS5000-E-(30-120k)-FM Datasheet",
    },
    # ── Delta Ultron HPH-120K ─────────────────────────────────────────────
    {
        "brand":             "Delta Electronics",
        "series":            "Ultron HPH",
        "model":             "HPH-120K",
        "topology":          "Double Conversion (Online)",
        "power_kw":          120,
        "power_kva":         120,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "40–70 Hz",
        "ac_in_current_a":   182,
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  182,
        "dc_nom_v":          480,           # ±240V bipolar = 480V between rails
        "dc_min_v":          384,           # 32 × 12V blocks EOD
        "dc_max_v":          552,           # 46 × 12V blocks max
        "dc_discharge_a":    250,           # 120000 ÷ 480
        "dc_charge_a":       40,            # Optional high-current charger
        "batt_type":         "VRLA (sealed lead-acid)",
        "batt_kwh_cabinet":  120,           # 40 × 12V × 250Ah = 120 kWh
        "batt_cab_v":        "480 V DC (40 × 12V, or 32–46 blocks configurable)",
        "eff_online_pct":    96,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1760 × 800 × 520",
        "ups_weight_kg":     312,
        "batt_cab_dims":     "1800 × 600 × 800 (external VRLA cabinet)",
        "batt_cab_weight_kg": 1100,
        "comms":             "SNMP (SNMP card option), RS-232, dry contacts, USB",
        "sg_presence":       "Medium — used in medium enterprise and edge DCs",
        "suitable_120kwh":   True,
        "notes":             "DC bus is ±240V bipolar architecture (split positive/negative). "
                             "Total positive-to-negative = 480V. Battery quantity: "
                             "32–46 × 12V blocks (configurable). Standard charger 20A; "
                             "optional 40A for faster recharge. No Li-ion option in base spec. "
                             "Source: [DE] deltapowersolutions.com HPH series spec page",
    },
    # ── Vertiv Liebert EXL S1 — 250 kW (smallest model; reference only) ────
    {
        "brand":             "Vertiv",
        "series":            "Liebert EXL S1",
        "model":             "EXL S1 250 kVA / 250 kW",
        "topology":          "Double Conversion (Online) — Transformer-free",
        "power_kw":          250,
        "power_kva":         250,
        "output_pf":         1.0,
        "ac_in_voltage_v":   "380 / 400 / 415 V (3-phase 4W+PE)",
        "ac_in_freq_hz":     "45–65 Hz",
        "ac_in_current_a":   380,           # 250000/(400×√3) ≈ 361A; nominal ~380A
        "ac_out_voltage_v":  "380 / 400 / 415 V (3-phase)",
        "ac_out_current_a":  361,
        "dc_nom_v":          540,           # Typical EXL S1 DC bus
        "dc_min_v":          432,           # EOD
        "dc_max_v":          612,           # Float charge (51 × 12V)
        "dc_discharge_a":    463,           # 250000 ÷ 540
        "dc_charge_a":       50,            # Standard
        "batt_type":         "VRLA / Li-ion LFP (Vertiv Li-ion battery system)",
        "batt_kwh_cabinet":  120,           # 45 × 12V × 222Ah ≈ 120 kWh
        "batt_cab_v":        "540 V DC (45 × 12V blocks typical)",
        "eff_online_pct":    96,
        "eff_eco_pct":       99,
        "ups_dims_hwdmm":    "1900 × 800 × 960",
        "ups_weight_kg":     720,
        "batt_cab_dims":     "1900 × 800 × 960 (Vertiv battery cabinet)",
        "batt_cab_weight_kg": 1200,
        "comms":             "SNMP (Vertiv Environet Alert), Modbus, BACnet, REPO",
        "sg_presence":       "Medium-High — large Tier 3/4 data centres",
        "suitable_120kwh":   False,
        "notes":             "EXL S1 starts at 250 kW — oversized for 120 kWh target "
                             "(gives only 29 min runtime). Included for reference as it is "
                             "deployed in Singapore hyperscale facilities. DC bus 540V is "
                             "higher than most peers — check battery string compatibility. "
                             "Source: [VT] Liebert EXL S1 500-600 kVA Datasheet",
    },
]

# ─── Column definitions ───────────────────────────────────────────────────────
COLUMNS = [
    # (header, field_key, width, format, notes_in_header)
    ("Brand",                          "brand",              22, None, ""),
    ("Series",                         "series",             18, None, ""),
    ("Model",                          "model",              26, None, ""),
    ("Topology",                       "topology",           28, None, ""),
    ("Power Rating\n(kW)",             "power_kw",           12, None, ""),
    ("Power Rating\n(kVA)",            "power_kva",          12, None, ""),
    ("Output Power\nFactor",           "output_pf",          10, None, ""),
    ("AC Input Voltage\n(V, phases)",  "ac_in_voltage_v",    28, None, ""),
    ("AC Input Freq\n(Hz)",            "ac_in_freq_hz",      14, None, ""),
    ("AC Input Current\n(A) @ rated",  "ac_in_current_a",    14, None, ""),
    ("AC Output\nVoltage (V)",         "ac_out_voltage_v",   24, None, ""),
    ("AC Output\nCurrent (A)",         "ac_out_current_a",   14, None, ""),
    ("DC Bus Voltage\nNominal (V)",    "dc_nom_v",           14, None, ""),
    ("DC Bus Voltage\nMin / EOD (V)",  "dc_min_v",           16, None, ""),
    ("DC Bus Voltage\nMax / Float (V)","dc_max_v",           16, None, ""),
    ("Max DC Discharge\nCurrent (A)",  "dc_discharge_a",     16, None, ""),
    ("Max DC Charge\nCurrent (A)",     "dc_charge_a",        14, None, ""),
    ("Battery Type",                   "batt_type",          40, None, ""),
    ("Battery Cabinet\nVoltage (VDC)", "batt_cab_v",         28, None, ""),
    ("Battery Energy\nper Cabinet (kWh)", "batt_kwh_cabinet",16, None, ""),
    ("Efficiency\nOnline Mode (%)",    "eff_online_pct",     14, None, ""),
    ("Efficiency\nECO Mode (%)",       "eff_eco_pct",        12, None, ""),
    ("UPS Dimensions\nH×W×D (mm)",    "ups_dims_hwdmm",     26, None, ""),
    ("UPS Weight\n(kg)",               "ups_weight_kg",      12, None, ""),
    ("Battery Cabinet\nDimensions",    "batt_cab_dims",      36, None, ""),
    ("Battery Cabinet\nWeight (kg)",   "batt_cab_weight_kg", 16, None, ""),
    ("Communication /\nIntegration",  "comms",              44, None, ""),
    ("SG DC Market\nPresence",         "sg_presence",        26, None, ""),
    ("Suitable for\n~120 kWh?",       "suitable_120kwh",    14, None, ""),
    ("Notes / Source",                 "notes",              60, None, ""),
]

def bool_to_str(v):
    return "YES ✔" if v else "NO — see notes"


def build_workbook():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: UPS Comparison ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "UPS Comparison"
    ws.freeze_panes = "C4"  # freeze brand + series, scroll right for specs

    # ── Title row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1,
        value="SINGAPORE DATA CENTRE UPS COMPARISON  —  Battery Integration Focus (~120 kWh)")
    title_cell.fill    = make_fill(C_TITLE_BG)
    title_cell.font    = Font(name="Calibri", size=14, bold=True, color=C_HEADER_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Subtitle row ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    sub_cell = ws.cell(row=2, column=1,
        value=(
            "Sources: Schneider Electric Galaxy VS IEC Spec (990-91141L_EN)  |  "
            "Eaton 93PM APAC Tech Spec Rev 2.653  |  "
            "Huawei UPS5000-E-(30-120k)-FM Datasheet  |  "
            "Delta Ultron HPH Spec Page (deltapowersolutions.com)  |  "
            "Vertiv Liebert EXL S1 Datasheet  —  Compiled by HyESys Agent, June 2026"
        ))
    sub_cell.fill      = make_fill(C_HEADER_BG)
    sub_cell.font      = Font(name="Calibri", size=9, italic=True, color="DDDDDD")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Column headers (row 3) ───────────────────────────────────────────────
    ws.row_dimensions[3].height = 50
    for col_idx, (hdr, _, _, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill      = make_fill(C_SUBHDR_BG)
        cell.font      = Font(name="Calibri", size=9, bold=True, color=C_SUBHDR_FG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border(color="FFFFFF")

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_i, udata in enumerate(UPS_DATA):
        r = row_i + 4
        is_suitable = udata.get("suitable_120kwh", False)
        is_120kw    = udata.get("power_kw") in (120,)
        row_bg = C_HIGHLIGHT if (is_suitable and is_120kw) else (
                 C_ALT_ROW  if row_i % 2 == 0 else C_WHITE)
        ws.row_dimensions[r].height = 40

        for col_idx, (_, field, _, _, _) in enumerate(COLUMNS, start=1):
            val = udata.get(field, "")
            if field == "suitable_120kwh":
                val = bool_to_str(val)
            cell = ws.cell(row=r, column=col_idx, value=val)

            # Background
            bg = row_bg
            if field == "dc_discharge_a" and isinstance(val, (int, float)) and val > 300:
                bg = C_WARN_BG   # flag high discharge currents
            if field == "suitable_120kwh" and "YES" in str(val):
                bg = C_GOOD_BG
            if field == "suitable_120kwh" and "NO" in str(val):
                bg = C_WARN_BG

            cell.fill      = make_fill(bg)
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = Alignment(
                horizontal="center" if isinstance(val, (int, float)) else "left",
                vertical="center",
                wrap_text=True)
            cell.border    = make_border()

    # ── Set column widths ────────────────────────────────────────────────────
    for col_idx, (_, _, width, _, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── Sheet 2: DC Battery Integration Notes ───────────────────────────────
    ws2 = wb.create_sheet("DC Integration Notes")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 70

    notes_data = [
        ("SECTION", "DETAIL"),
        ("PURPOSE OF THIS SHEET", "Key engineering notes for integrating external battery storage (e.g., HyESys HySBatt) with data centre UPS DC buses."),
        ("", ""),
        ("DC BUS VOLTAGE SUMMARY", ""),
        ("Schneider Galaxy VS",      "480 V DC nominal  |  EOD: 384 V  |  Float: 552–576 V"),
        ("Eaton 93PM",               "480 V DC nominal  |  EOD: 432 V  |  Float: 497–564 V"),
        ("Huawei UPS5000-E",         "480 V DC nominal  |  Range: 320–576 V (wide — Li-ion ready)"),
        ("Delta Ultron HPH",         "±240 V bipolar (=480 V total)  |  EOD: 384 V  |  Max: 552 V"),
        ("Vertiv Liebert EXL S1",    "540 V DC nominal  |  EOD: 432 V  |  Float: 612 V  ⚠ HIGHER VOLTAGE"),
        ("", ""),
        ("120 kWh SIZING NOTE",      ""),
        ("VRLA @ 480V",              "40 × 12V × 250Ah = 120 kWh  |  Approx 1,100 kg per cabinet"),
        ("VRLA @ 540V",              "45 × 12V × 222Ah = 120 kWh  |  Approx 1,200 kg per cabinet"),
        ("Li-ion LFP @ 480V",        "Huawei SmartLi: 10 × 4.8 kWh modules = 48 kWh per rack  →  3 racks = 144 kWh  |  ~720 kg total"),
        ("", ""),
        ("DC DISCHARGE CURRENT",     "At 480V DC, discharge current = Load_kW × 1000 / 480"),
        ("",                         "100 kW  →  208 A      120 kW  →  250 A      200 kW  →  417 A"),
        ("",                         "⚠ High current (>300A) requires large-gauge DC cable and robust contactors."),
        ("", ""),
        ("CHARGE CURRENT NOTE",      "Standard UPS chargers are conservative (20–50A) to prolong VRLA life."),
        ("",                         "For HyESys/HySBatt integration, confirm UPS charger won't conflict with external BMS charge control."),
        ("",                         "Preferred approach: HyESys controls DC side via BMS; UPS charger set to float-only mode."),
        ("", ""),
        ("COMMUNICATION",            "All listed UPS models support SNMP and Modbus TCP for remote monitoring."),
        ("",                         "Dry-contact relay outputs provide: On Battery / Low Battery / Fault / EPO signals."),
        ("",                         "For HyESys integration, use Modbus TCP registers for real-time voltage, current, SoC."),
        ("", ""),
        ("VRLA vs LI-ION DECISION",  ""),
        ("VRLA Pros",                "Lower upfront cost | Widely available in SG | No BMS integration complexity"),
        ("VRLA Cons",                "Heavy (~1,100 kg/cabinet) | ~500 cycle life | Requires quarterly checks"),
        ("Li-ion (LFP) Pros",        "3,000+ cycles | Lighter (~720 kg for 144 kWh) | Better DC range compatibility"),
        ("Li-ion (LFP) Cons",        "Higher upfront cost | Requires BMS protocol matching with UPS"),
        ("", ""),
        ("VERTIV EXL S1 CAUTION",    "⚠ DC bus 540V is 60V higher than Galaxy VS / Eaton 93PM / Huawei UPS5000-E."),
        ("",                         "HySBatt packs must support 540V+ range if integrating with EXL S1."),
        ("",                         "Verify HySBatt max DC voltage before connecting."),
        ("", ""),
        ("RECOMMENDED MODELS",       "For ~120 kWh battery integration in Singapore DCs:"),
        ("1st Choice",               "Schneider Galaxy VS 120 kW (GVSUPS120KHS)  —  most common, best SG support"),
        ("2nd Choice",               "Eaton 93PM 100 kW  —  modular, Li-ion capable, high efficiency (97%)"),
        ("3rd Choice",               "Huawei UPS5000-E 120K-FM  —  widest DC range (320–576V), modular, SmartLi ready"),
        ("Not Recommended",          "Vertiv EXL S1 — oversized for 120 kWh, higher DC voltage, higher cost"),
        ("", ""),
        ("DATA SOURCES",             ""),
        ("[SE]",  "Schneider Electric Galaxy VS IEC Tech Spec 990-91141L_EN  (productinfo.se.com)"),
        ("[EA]",  "Eaton 93PM APAC Tech Spec Rev 2.653  (eaton.com/content/dam/...)"),
        ("[HW]",  "Huawei UPS5000-E-(30-120k)-FM Datasheet  (digitalpower.huawei.com)"),
        ("[DE]",  "Delta Ultron HPH Series Specifications  (deltapowersolutions.com)"),
        ("[VT]",  "Vertiv Liebert EXL S1 500-600 kVA Datasheet  (vertiv.com)"),
        ("",      "Compiled by HyESys Agent, June 2026"),
    ]

    # Row 1: header
    for ri, (col_a, col_b) in enumerate(notes_data, start=1):
        ca = ws2.cell(row=ri, column=1, value=col_a)
        cb = ws2.cell(row=ri, column=2, value=col_b)
        ws2.row_dimensions[ri].height = 18

        if col_a == "SECTION":
            for c in [ca, cb]:
                c.fill = make_fill(C_HEADER_BG)
                c.font = Font(name="Calibri", size=10, bold=True, color=C_HEADER_FG)
        elif col_a in ("DC BUS VOLTAGE SUMMARY", "120 kWh SIZING NOTE",
                       "DC DISCHARGE CURRENT", "CHARGE CURRENT NOTE",
                       "COMMUNICATION", "VRLA vs LI-ION DECISION",
                       "VERTIV EXL S1 CAUTION", "RECOMMENDED MODELS",
                       "DATA SOURCES", "PURPOSE OF THIS SHEET"):
            ca.fill = make_fill(C_SUBHDR_BG)
            ca.font = Font(name="Calibri", size=10, bold=True, color=C_SUBHDR_FG)
            cb.fill = make_fill(C_SUBHDR_BG)
            cb.font = Font(name="Calibri", size=10, bold=True, color=C_SUBHDR_FG)
        elif col_a in ("1st Choice", "2nd Choice", "3rd Choice"):
            ca.fill = make_fill(C_GOOD_BG)
            cb.fill = make_fill(C_GOOD_BG)
            ca.font = Font(name="Calibri", size=9, bold=True)
            cb.font = Font(name="Calibri", size=9)
        elif col_a in ("Not Recommended", "VRLA Cons", "Li-ion (LFP) Cons"):
            ca.fill = make_fill(C_WARN_BG)
            cb.fill = make_fill(C_WARN_BG)
            ca.font = Font(name="Calibri", size=9)
            cb.font = Font(name="Calibri", size=9)
        else:
            bg = C_SECTION_BG if ri % 2 == 0 else C_WHITE
            ca.fill = make_fill(bg)
            cb.fill = make_fill(bg)
            ca.font = Font(name="Calibri", size=9,
                           bold=(col_b != "" and col_a != ""))
            cb.font = Font(name="Calibri", size=9)

        ca.alignment = Alignment(horizontal="left", vertical="center")
        cb.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ca.border = make_border()
        cb.border = make_border()

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()
