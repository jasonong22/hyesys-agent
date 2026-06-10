import sys, math, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Singapore DC Battery Cell Comparison.xlsx')
ws = wb['Cell Comparison']
headers = [cell.value for cell in ws[1]]

model_col  = headers.index('Model')
energy_col = headers.index('Energy (Wh)')
weight_col = headers.index('Weight (g)')
dim_col    = headers.index('Dimensions (mm)')
whkg_col   = headers.index('Wh/kg')
whm3_col   = headers.index('Wh/m3')

def parse_volume_cm3(dim_str):
    s = str(dim_str).strip()
    if 'x H' in s or re.match(r'[Ø]([\d.]+)\s*x', s):
        m = re.search(r'([\d.]+)\s*[xX]\s*H?([\d.]+)', s)
        if m:
            d_mm, h_mm = float(m.group(1)), float(m.group(2))
            return math.pi * (d_mm / 2) ** 2 * h_mm / 1000
    nums = [float(x) for x in re.findall(r'[\d.]+', s) if float(x) > 1]
    if len(nums) >= 3:
        return nums[0] * nums[1] * nums[2] / 1000
    return None

issues = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    model = row[model_col]
    if not model:
        continue
    energy_raw = row[energy_col]
    weight_raw = row[weight_col]
    dim        = row[dim_col]
    whkg_stored = row[whkg_col]
    whm3_stored = row[whm3_col]
    if not energy_raw:
        continue
    energy = float(str(energy_raw).replace(',', ''))

    # Wh/kg
    weight_g = None
    if weight_raw is not None:
        try:
            weight_g = float(str(weight_raw).replace('g', '').replace(',', '').strip())
        except Exception:
            pass
    whkg_calc = round(energy / (weight_g / 1000)) if weight_g else None
    if whkg_stored is not None and whkg_calc is not None:
        kg_ok = (whkg_stored == whkg_calc)
    else:
        kg_ok = None

    # Wh/m3
    vol_cm3 = parse_volume_cm3(dim) if dim else None
    whm3_calc = round(energy / (vol_cm3 * 1e-6)) if vol_cm3 else None
    if whm3_stored is not None and whm3_calc is not None:
        m3_ok = (whm3_stored == whm3_calc)
    else:
        m3_ok = None

    vol_str = str(round(vol_cm3, 2)) if vol_cm3 else '-'

    print(f"Model: {model}")
    print(f"  Energy={energy} Wh  |  Weight={weight_raw}  ->  Wh/kg stored={whkg_stored}  calc={whkg_calc}  {'OK' if kg_ok else ('MISMATCH' if kg_ok is False else 'N/A')}")
    print(f"  Dims={dim}  |  Vol={vol_str} cm3  ->  Wh/m3 stored={whm3_stored}  calc={whm3_calc}  {'OK' if m3_ok else ('MISMATCH' if m3_ok is False else 'N/A')}")
    print()

    if kg_ok is False:
        issues.append(f"Wh/kg MISMATCH: {model} stored={whkg_stored} calc={whkg_calc}")
    if m3_ok is False:
        issues.append(f"Wh/m3 MISMATCH: {model} stored={whm3_stored} calc={whm3_calc}")

if issues:
    print("ISSUES FOUND:")
    for i in issues:
        print(" ", i)
else:
    print("All values match.")
