import sys, math
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Singapore DC Battery Cell Comparison.xlsx')
ws = wb['Cell Comparison']
headers = [cell.value for cell in ws[1]]

celltype_col = headers.index('Cell Type')
whm3_col     = headers.index('kWh/m3')
whkg_col     = headers.index('Wh/kg')
model_col    = headers.index('Model')

cyl_whm3, pri_whm3 = [], []
print(f"{'Cell Type':<25} {'Model':<30} {'kWh/m3':>10} {'Wh/kg':>8}")
print('-'*75)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[model_col]:
        ct = str(row[celltype_col] or '')
        print(f"{ct:<25} {str(row[model_col]):<30} {str(row[whm3_col]):>10} {str(row[whkg_col]):>8}")
        if row[whm3_col]:
            if 'Cyl' in ct:
                cyl_whm3.append(row[whm3_col])
            elif 'Pris' in ct:
                pri_whm3.append(row[whm3_col])

print()
print(f"Cylindrical kWh/m3 range: {min(cyl_whm3):.1f} – {max(cyl_whm3):.1f}")
print(f"Prismatic   kWh/m3 range: {min(pri_whm3):.1f} – {max(pri_whm3):.1f}")

# Packing geometry check
# Hexagonal close packing efficiency = pi/(2*sqrt(3)) = 90.69%
hex_eff = math.pi / (2 * math.sqrt(3))
sq_eff  = math.pi / 4
print()
print(f"Cylindrical packing efficiency (hex): {hex_eff*100:.1f}%  -> {(1-hex_eff)*100:.1f}% waste")
print(f"Cylindrical packing efficiency (sq):  {sq_eff*100:.1f}%  -> {(1-sq_eff)*100:.1f}% waste")
print("Practical module-level (with cooling): ~65-80%  -> 20-35% waste")
print()
print("Pack-level kWh/m3 estimates:")
for eff, label in [(0.70, 'cyl 70%'), (0.75, 'cyl 75%'), (0.85, 'pri 85%'), (0.92, 'pri 92%')]:
    if 'cyl' in label:
        lo = min(cyl_whm3) * eff
        hi = max(cyl_whm3) * eff
    else:
        lo = min(pri_whm3) * eff
        hi = max(pri_whm3) * eff
    print(f"  {label}: cell range × {eff} = {lo:.0f} – {hi:.0f} kWh/m3 pack-level")
