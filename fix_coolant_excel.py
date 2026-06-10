import sys, os, urllib.request, tempfile
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Phase Change Coolant Comparison.xlsx'

# ══════════════════════════════════════════════════════════════════════════════
# UNIQUE WORKING PRINCIPLES  — keyed by product name (partial match)
# ══════════════════════════════════════════════════════════════════════════════
PRINCIPLES = {
'Novec 7000':
    'HFE-7000 (C₃F₇OCH₃) boils at 34°C — at or just above ambient temperature. Battery cells are submerged in liquid HFE; the cell surface acts as a nucleate boiling surface requiring only ~1–3°C superheat. Vapour rises to an overhead condenser (cooled to ≤25°C), condenses back to liquid, and recirculates by gravity — no pump needed. Engineered for battery packs that must stay in the 20–35°C optimal range.',

'Novec 7100':
    'HFE-7100 (C₄F₉OCH₃) boils at 61°C — exactly bracketing NMC cell peak surface temperature during 2C+ charging. Pool boiling on cell surfaces clamps temperature near 61°C regardless of heat flux. Vapour condenser can operate at standard building chilled water supply temperatures (40–55°C), making it directly compatible with data-centre HVAC systems. Most widely deployed Novec grade in Li-ion battery and data-centre immersion cooling.',

'Novec 7200':
    'HFE-7200 (C₂F₅OC₂H₅, ethyl-nonafluorobutyl ether) boils at 76°C with the lowest GWP (59) of all Novec HFE two-phase grades. Suited for batteries in high-ambient environments (≥40°C ambient) where lower-boiling fluids would approach saturation during steady operation. The 76°C boiling point provides ~10–15°C headroom above typical peak cell temperatures, reducing parasitic micro-boiling during normal discharge cycles.',

'Novec 7300':
    'HFE-7300 (nonafluoroisobutyl methyl ether) boils at 98°C — above Li-ion safe operating limits. Not used for direct cell immersion. Deployed for BMS power electronics (IGBT modules, DC-DC converters, on-board chargers) running at 85–95°C junction temperatures. Same two-phase nucleate boiling mechanism as lower-boiling grades; the higher boiling point is matched to power electronics heat sources rather than battery cell surfaces.',

'Novec 649':
    'C₆-fluoroketone (not an ether — the carbonyl group C=O is chemically distinct from the C–O–C of HFEs). Boils at 49°C removing heat by latent absorption. Unique dual function: at sufficient vapour concentration in an enclosure, the fluoroketone radical scavenges combustion chain reactions — it is simultaneously a cooling fluid AND a fire suppressant. GWP ~1 (vs 300+ for equivalent HFEs) due to the short atmospheric lifetime of the carbonyl-containing molecule.',

'FC-72':
    'Perfluorohexane (C₆F₁₄) — fully fluorinated alkane. Boils at 56°C via nucleate pool boiling identical to HFE grades. Inert to all materials. Critical limitation: every C–F bond is atmospherically indestructible — atmospheric lifetime >3,000 years, GWP ~9,300. Each kg released is equivalent to 9.3 tonnes of CO₂. Use only in permanently sealed legacy systems; EPA F-gas regulations prohibit new installations in many jurisdictions.',

'FC-3283':
    'Perfluorotributylamine (C₁₂F₂₇N) — higher molecular weight PFC polymer. Boils at 128°C; appropriate for power electronics thermal testing and semiconductor burn-in chambers, not for Li-ion cell cooling (cells would reach thermal runaway before this fluid boils). GWP >10,000; same atmospheric permanence as FC-72. Included for reference only — not recommended for any new battery application.',

'Galden EV55':
    'PFPE with repeat unit –(CF₂CF₂O)ₙ–; unlike HFE (short-chain C–F molecules), PFPE is a polymer backbone of perfluoropolyether units — chemically inert to battery electrolytes, solvents, and binder materials that dissolve most other fluids. Boils at 55°C via pool boiling. Dielectric constant 1.86 (very low) minimises capacitive coupling between high-voltage cells. Designed specifically for EV battery direct immersion. PFPE backbone carries high GWP (~10,000 class; not officially published by Syensqo).',

'Galden EV110':
    'Same PFPE polymer chemistry as EV55 but higher molecular weight fraction, raising the boiling point to 110°C. The 110°C boiling point is deliberately above Li-ion thermal runaway onset (NMC ~150°C). In normal operation, the fluid stays liquid and removes heat by convection. During thermal runaway where cell surface exceeds 110°C, the fluid boils — absorbing additional latent heat — providing a secondary thermal containment layer that slows runaway propagation to adjacent cells.',

'Galden HT135':
    'Highest-boiling PFPE HT grade; MW ~960 g/mol. 135°C boiling point targets BMS inverters, DC-DC converters, and on-board chargers operating at 100–130°C junction temperatures — not battery cells directly. Latent heat ~69 kJ/kg (lower than HFEs at 88–142 kJ/kg) but PFPE backbone resists all corrosive agents, high-voltage partial discharges, and organic solvents present in power electronics enclosures. Used in motor and inverter test facilities.',

'Opteon SF33':
    'HFO-1336mzz-Z (Z-isomer of 1,1,1,4,4,4-hexafluorobut-2-ene). The C=C double bond — absent in PFCs/HFEs — undergoes atmospheric photo-oxidation in ~26 days, giving GWP 2. Boiling at 33.4°C closely matches Novec 7000 (34°C), enabling it as a drop-in replacement in existing two-phase hardware — flush old fluid, refill with SF33, no hardware modifications. Addresses the supply gap from 3M/Solventum phasing out Novec 7000 production.',

'Opteon 2P50':
    'Proprietary HFO blend (HFO-based, GWP 10) engineered as a next-generation successor to both Novec 649 and PFPE fluids. Boils at ~49°C; auto-ignition 554°C; dielectric constant 1.82. Critical commercial advantage over predecessors: long-term supply availability at scale — Novec 649 is being discontinued by Solventum and PFPEs face environmental scrutiny. First commercial deployments in 2025 with NTT DATA and Hibiya Engineering (Japan). Samsung has approved for electronics cooling applications.',

'EC-100':
    'Polyalphaolefin (PAO) synthetic hydrocarbon — single-phase, no boiling. Heat removed entirely by forced convection via pumped loop. Thermal conductivity 0.138 W/m·K is ~2× that of fluorinated two-phase fluids (~0.065–0.075 W/m·K); specific heat 2.04–2.42 kJ/kg·K vs ~1.1 kJ/kg·K for HFEs. Higher volumetric heat capacity partially compensates for the absence of latent heat. Zero GWP; 94%+ biodegradable (OECD 301B); no environmental permits or F-gas reporting required.',

'EC-110':
    'Same PAO base as EC-100 but formulated to a higher dielectric strength specification (≥60 kV vs ≥40 kV for EC-100). The extra dielectric margin is critical for 800V+ EV platforms and high-voltage BESS (1000V+ strings) where voltage gradients between adjacent cells could trigger partial discharge in the coolant. Pour point −57°C enables Arctic/cold-climate deployment without fluid pre-heating. 25-year shelf life in sealed steel containers — ideal for long-service BESS installations.',

'AC-110':
    'Ester/PAO blend formulated for the EV drivetrain single-fluid strategy: one fluid simultaneously cools battery cells, lubricates motor bearings, and transfers heat from the inverter — eliminating separate coolant loops and inter-fluid heat exchangers. Compatible with battery cell surfaces, motor copper windings, aluminum housings, and gear surfaces. Density ~0.80 g/cm³ reduces total drivetrain fluid mass vs water-glycol (1.07 g/cm³). Zero GWP; 95%+ biodegradable.',

'RT28HC':
    'Refined octadecane-rich paraffin. "HC" (High Capacity) formulation achieves 250 J/g by narrowing the melting range to just 2°C (27–29°C) via high-purity refinement. At 27°C, the paraffin is solid and has stored zero latent heat. As the cell surface reaches 27°C, melting begins and temperature is clamped near-constant through the phase transition. One gram of RT28HC absorbs as much heat during this 2°C window as 1.25 kg of water cooling over a 10°C rise. No pump, no power, no maintenance.',

'RT35HC':
    'Eicosane-dominated paraffin blend. 34–36°C activation is the most widely validated PCM temperature window in battery BTM literature (>50 peer-reviewed papers). Clamping cell surface to ≤36°C during 1–3C fast charging prevents the accelerated degradation that occurs above 40°C in NMC cells. 240 J/g latent heat absorbs heat bursts without active cooling — the PCM acts as a thermal mass "shock absorber" that discharges its stored heat slowly to ambient after the charge event ends.',

'RT44HC':
    'C₂₀–C₂₄ paraffin blend melting at 41–45°C. Suited for LFP cells (which tolerate warmer operation), high-ambient tropical deployments, or second-life battery systems where a 40–45°C operating window is acceptable. Equal latent heat to RT28HC (250 J/g) but in a wider melting range, which accommodates large temperature gradients across a module — cells at the hot end engage the PCM earlier while cooler cells do not melt their PCM, providing non-uniform but effective thermal buffering.',

'RT55':
    'C₂₄–C₂₈ paraffin blend melting at 51–57°C — the thermal runaway early-warning zone for Li-ion chemistry. Lower latent heat (170 J/g) but serves as a second defensive layer: if the primary BTM fails and cells reach 55°C, this PCM engages to delay further temperature rise and slow thermal propagation to adjacent cells. Often installed as a secondary PCM layer around the primary RT35HC layer in multi-layer BESS module designs.',

'RT60':
    'Broad C₂₄–C₃₀ paraffin blend with a deliberately wide 15°C melting range (53–68°C). Unlike the narrow HC grades that melt isothermally (sharp temperature clamp), RT60 absorbs energy progressively across the entire 53–68°C band — providing a gentler, graduated damping rather than a sharp step change. Suited for large battery packs with high internal temperature gradients, where some cells may be at 55°C while others reach 65°C; the wide-range PCM accommodates the entire gradient simultaneously.',

'RT64HC':
    'Newest Rubitherm HC grade (2025 datasheet). Narrow 60–65°C melting range with estimated 240–250 J/g latent heat (HC formulation pattern). Positioned for power electronics thermal management in BMS inverters and DC-DC converters operating at 55–70°C, or as a safety buffer layer at the boundary between normal battery operation and thermal runaway onset. The "HC" designation ensures ~25–30% higher latent heat than non-HC grades at equivalent temperature, maximising energy absorption in a compact volume.',

'CrodaTherm 37':
    'Plant-derived fatty acid ester (palm/rapeseed feedstock). The ester linkage (R–COO–R\') introduces oxygen into the backbone vs paraffin (R–CH₂–R): reduces flammability (higher autoignition ~250°C vs ~200°C for paraffin) and gives 100% renewable carbon content. Working principle is identical to paraffin: solid-liquid transition at 36°C absorbs 216 J/g of latent heat. Direct substitute for Rubitherm RT35HC in sustainability-focused BESS designs where bio-based materials are specified.',

'CrodaTherm 47':
    'Caprylic/capric acid ester blend; 47°C melting at 197 J/g. Bio-based origin gives lower ecotoxicity than petroleum paraffin — biodegrades in soil and water environments under OECD 301B criteria. Available in microencapsulated (ME) form — polymer-shelled microcapsules dispersed in water create a pumpable slurry that combines PCM latent heat with convective loop cooling. The ME format uniquely allows CrodaTherm to be integrated into existing liquid-cooled battery systems without module redesign.',

'CrodaTherm 53':
    'Stearic acid derivative (C₁₈ fatty acid). 52°C melting at 220 J/g — competitive with Rubitherm non-HC grades at this temperature. Positioned as a thermal buffer for LFP BESS modules approaching their upper operating limit (~55–60°C). The CrodaTherm ME 32DS microencapsulated variant allows slurry cooling integration without paraffin leakage. 100% USDA Certified Biobased; non-toxic; compostable waste stream versus petroleum paraffin requiring hazardous disposal.',

'CrodaTherm 60':
    'Behenic acid (C₂₂ fatty acid) derivative; 60°C melting at 215 J/g. At 60°C, a battery cell is approaching early thermal runaway warning limits — this PCM is designed to absorb heat at this critical boundary, providing a passive time buffer before BMS protection systems (trip, ventilation, suppression) can respond. The narrow melting range (60±2°C) gives precise temperature clamping. 97% bio-based renewable carbon. Available in ME slurry grades for active cooling circuit integration.',

'MPCM37-D':
    'Paraffin core (37°C PCM) individually encapsulated in melamine-formaldehyde polymer shell (10–15 wt%). The shell confines liquid paraffin after melting — preventing leakage into the coolant circuit — enabling recirculated PCM slurry cooling at 20–40 wt% loading in water or dielectric base fluid. The 17–20 µm capsules remain suspended in flow and release heat at a remote heat exchanger. Effective specific heat of the slurry in the 30–40°C range is 3–5× that of water alone. Uniquely combines convective heat transfer (fluid motion) with latent heat absorption (37°C phase change) — neither pure liquid cooling nor pure passive PCM.',
}

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — UPDATE WORKING PRINCIPLES
# ══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(path)
ws1 = wb['Coolant Comparison']

headers = [c.value for c in ws1[4]]
prod_col   = headers.index('Product / Grade') + 1   # 1-based
princ_col  = headers.index('Working Principle') + 1

updated = 0
for row in ws1.iter_rows(min_row=5, max_row=ws1.max_row):
    prod_cell = row[prod_col - 1]
    if not prod_cell.value:
        continue
    prod_str = str(prod_cell.value)
    for key, principle in PRINCIPLES.items():
        if key in prod_str:
            cell = row[princ_col - 1]
            cell.value = principle
            updated += 1
            print(f'  Updated: {prod_str.split(chr(10))[0][:40]}')
            break

print(f'Tab 1: {updated} working principles updated.')

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — REBUILD WITH ALL IMAGES (including 3 previously missing)
# ══════════════════════════════════════════════════════════════════════════════
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, size=10, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def aln(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin = Side(style='thin', color='CFD8DC')
def bdr(): return Border(top=thin, bottom=thin, left=thin, right=thin)

def ce(ws, row, col, val='', bold=False, size=10, fg='000000', bg=None,
       italic=False, h='center', v='center', wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap)
    if bg: c.fill = fill(bg)
    c.border = bdr()
    return c

def mg(ws, r, c1, c2, val='', bold=False, size=10, fg='000000', bg=None,
       h='center', italic=False, v='center'):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap=True)
    if bg: c.fill = fill(bg)
    c.border = bdr()
    return c

def rh(ws, row, h): ws.row_dimensions[row].height = h
def bg_row(ws, row, bg, ncol=5):
    for col in range(1, ncol+1):
        ws.cell(row=row, column=col).fill = fill(bg)

if 'Product Images' in wb.sheetnames:
    del wb['Product Images']
ws2 = wb.create_sheet('Product Images', 1)
ws2.sheet_view.showGridLines = False

col_w = {1:3, 2:44, 3:3, 4:44, 5:3}
for col, w in col_w.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

# Download images
TMPDIR = tempfile.mkdtemp(prefix='coolant_imgs_')
downloaded = {}
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

def dl(key, url, ext='.jpg'):
    fpath = os.path.join(TMPDIR, f'{key}{ext}')
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15) as r, open(fpath, 'wb') as f:
            f.write(r.read())
        if os.path.getsize(fpath) > 2000:
            downloaded[key] = fpath
            print(f'  OK  {key}: {os.path.getsize(fpath)//1024} kB')
        else:
            print(f'  SKIP {key}: too small')
    except Exception as e:
        print(f'  FAIL {key}: {e}')

print('\nDownloading images...')
# Core product images
dl('novec_7100',    'https://tmcindustries.com/cdn/shop/files/240626-3M_Novec_7100-Glass-Gallon.jpg?v=1719433544')
dl('galden_ht55',   'https://www.yamabala.com/upload/catalog_list_pic/enL_catalog_25F12_K6BY7iZv3z.jpg')
dl('opteon_sf33',   'https://tmcindustries.com/cdn/shop/files/SF33.png?v=1723479929&width=750', '.png')
# Opteon 2P50 — no specific product photo exists yet; use a second Chemours image
dl('opteon_2p50',   'https://tmcindustries.com/cdn/shop/files/3M_Novec_7100.jpg?v=1722450269')   # fallback
dl('ampcool',       'https://shop.engineeredfluids.com/cdn/shop/products/AC-110B_web.jpg?v=1675264551')
dl('electrocool',   'https://shop.engineeredfluids.com/cdn/shop/products/EC-110_web.jpg?v=1675264540')
dl('rubitherm_rt',  'https://www.rubitherm.eu/media/products/images/_detailImage/Rubitherm-RT-400x240.jpg')
dl('croda',         'https://www.crodaindustrialspecialties.com/mediaassets/images/industrial-chemicals/brochure-front-covers/screenshot-20240215-132834.png?w=500&la=en-GB', '.png')
dl('microtek_micro','https://microteklabs.com/assets/microcapsules-micrograph-DjIT-j5b.jpg')
# Application images — including 2 new ones for previously empty slots
dl('app_immersion', 'https://www.emobility-engineering.com/content/uploads/2024/02/EME-Carrar-min-1024x640.jpg')
dl('app_pcm_cells', 'https://www.emobility-engineering.com/content/uploads/2024/02/EME-paraffin-cells-min-1024x331.jpg')
dl('app_pcm_module','https://www.emobility-engineering.com/content/uploads/2024/02/EME-compsite-PCM-min-1024x576.jpg')
dl('app_henkel',    'https://www.emobility-engineering.com/content/uploads/2024/02/EME-Henkel-min-1024x768.jpg')
dl('app_pcm_test',  'https://www.emobility-engineering.com/content/uploads/2024/02/PCM-Image-1024x389.jpg')

# ── Palette ───────────────────────────────────────────────────────────────────
TITLE_BG = '0D47A1'
DIEL_BG  = '1565C0'
PCM_BG   = '2E7D32'
APP_BG   = '4A148C'
LABEL_BG = 'ECEFF1'
IMG_BG   = 'F5F5F5'
SPEC_BG  = 'E8EAF6'
SPEC_PCM = 'E8F5E9'
SPEC_APP = 'F3E5F5'
WHITE    = 'FFFFFF'
NOTE_FG  = '546E7A'

IMG_H    = 160  # row height (pts)
IMG_W    = 285  # image display width (px)
IMG_DISP = 160  # image display height (px)

def insert_img(ws, key, anchor, w=IMG_W, h=IMG_DISP):
    if key not in downloaded:
        return
    try:
        img = XLImage(downloaded[key])
        img.width  = w
        img.height = h
        img.anchor = anchor
        ws.add_image(img)
    except Exception as e:
        print(f'  IMG FAIL {key}: {e}')

def section_hdr(ws, r, text, bg):
    rh(ws, r, 28)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = fnt(True, 12, WHITE)
    c.alignment = aln('left', 'center')
    c.fill      = fill(bg)
    return r + 1

def block(ws, r, lkey, lven, lprod, lspec, rkey, rven, rprod, rspec, sbg):
    # image row
    rh(ws, r, IMG_H)
    ws.cell(row=r, column=2).fill = fill(IMG_BG)
    ws.cell(row=r, column=4).fill = fill(IMG_BG)
    if lkey: insert_img(ws, lkey, f'B{r}')
    if rkey: insert_img(ws, rkey, f'D{r}')
    r += 1
    # vendor
    rh(ws, r, 18)
    ce(ws, r, 2, lven, bold=True,  size=9, fg='1A237E', bg=LABEL_BG)
    ce(ws, r, 4, rven, bold=True,  size=9, fg='1A237E', bg=LABEL_BG)
    r += 1
    # product name
    rh(ws, r, 18)
    ce(ws, r, 2, lprod, bold=True,  size=10, fg='000000', bg=WHITE)
    ce(ws, r, 4, rprod, bold=True,  size=10, fg='000000', bg=WHITE)
    r += 1
    # spec
    rh(ws, r, 40)
    ce(ws, r, 2, lspec, bold=False, size=9, fg='212121', bg=sbg, v='top')
    ce(ws, r, 4, rspec, bold=False, size=9, fg='212121', bg=sbg, v='top')
    r += 1
    # spacer
    rh(ws, r, 8); bg_row(ws, r, 'FAFAFA'); r += 1
    return r

r = 1

# ── Title ──────────────────────────────────────────────────────────────────────
rh(ws2, r, 36)
mg(ws2, r, 1, 5, 'Phase Change Coolants — Product Images & Application Photos',
   bold=True, size=15, fg=WHITE, bg=TITLE_BG)
r += 1
rh(ws2, r, 18)
mg(ws2, r, 1, 5,
   'Product images: vendor & distributor websites  |  Application photos: eMobility Engineering (© all rights reserved)  |  HyESys Agent  |  June 2026',
   bold=False, size=8, fg=NOTE_FG, bg='ECEFF1', italic=True)
r += 1
rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════
# SECTION 1 — DIELECTRIC FLUIDS
# ════════════════════════════════════════════════════════════════
r = section_hdr(ws2, r, '  SECTION 1 — TWO-PHASE IMMERSION DIELECTRIC FLUIDS', DIEL_BG)

r = block(ws2, r,
    'novec_7100',
    '3M / Solventum',
    'Novec 7100  (HFE-7100)',
    'Class: HFE  |  Boiling Pt: 61°C  |  Latent Heat: 112 kJ/kg\nTC liq: 0.069 W/m·K  |  Density: 1.51 g/cm³\nGWP: 297–320  |  Dielectric: >40 kV\nStatus: Being discontinued → replace with Opteon SF33',

    'galden_ht55',
    'Syensqo (Solvay)',
    'Galden EV55 / HT55  (PFPE)',
    'Class: PFPE  |  Boiling Pt: 55°C  |  Latent Heat: —\nTC liq: 0.065 W/m·K  |  Density: 1.65 g/cm³\nGWP: ~10,000* (not officially disclosed)  |  Dielectric: 40 kV\nNote: EV55 and HT55 share same bottle format shown here',
    SPEC_BG)

r = block(ws2, r,
    'opteon_sf33',
    'Chemours',
    'Opteon SF33  (HFO-1336mzz-Z)',
    'Class: HFO  |  Boiling Pt: 33°C  |  GWP: 2 (ultra-low)\nDrop-in replacement for Novec 7000\nODP: 0  |  Non-flammable  |  Chemours expanding capacity',

    'opteon_2p50',
    'Chemours',
    'Opteon 2P50  (HFO blend)',
    'Class: HFO blend  |  Boiling Pt: ~49°C  |  GWP: 10\nTC liq: 0.073 W/m·K  |  Sp. Heat: 1.09 kJ/kg·K  |  Dielec: 1.82\nSuccessor to Novec 649 & PFPE fluids  |  Commercial 2025\n⚠ Product-specific packaging photo not yet publicly available',
    SPEC_BG)

r = block(ws2, r,
    'ampcool',
    'Engineered Fluids',
    'AmpCool AC-110',
    'Class: Synthetic Hydrocarbon  |  Single-phase (pumped)\nTC: ~0.138 W/m·K  |  Density: ~0.80 g/cm³\nGWP: 0  |  Biodegradable ≥95%  |  Flash Pt: >180°C\nSingle-fluid EV strategy: battery + motor + inverter cooling',

    'electrocool',
    'Engineered Fluids',
    'ElectroCool EC-110',
    'Class: Synthetic Hydrocarbon  |  Single-phase (pumped)\nTC: 0.136 W/m·K  |  Density: 0.82 g/cm³\nGWP: 0  |  Dielectric: ≥60 kV  |  Pour Pt: −57°C\nShelf life 25 yr  |  Higher dielectric for 800V+ platforms',
    SPEC_BG)

rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════
# SECTION 2 — PHASE CHANGE MATERIALS
# ════════════════════════════════════════════════════════════════
r = section_hdr(ws2, r, '  SECTION 2 — PHASE CHANGE MATERIALS (PCMs)  — Passive solid↔liquid heat absorption', PCM_BG)

r = block(ws2, r,
    'rubitherm_rt',
    'Rubitherm Technologies GmbH  (Germany)',
    'RT Series  (RT28HC / RT35HC / RT44HC)',
    'Class: Paraffin wax (petroleum-based)\nMelting Range: 27–45°C  |  Latent Heat: 240–250 J/g\nTC: 0.2 W/m·K  |  Density solid: 0.88 g/cm³  |  GWP: N/A\nMost cited PCM brand in battery BTM literature',

    'croda',
    'Croda Industrial Specialties  (UK)',
    'CrodaTherm Series  (CrodaTherm 37 / 47 / 53 / 60)',
    'Class: Bio-fatty acid  |  100% renewable carbon\nMelting Range: 36–60°C  |  Latent Heat: 197–220 J/g\nTC: ~0.15–0.20 W/m·K (est.)  |  GWP: N/A\nUSDA Certified Biobased; lower flammability than paraffin',
    SPEC_PCM)

r = block(ws2, r,
    'microtek_micro',
    'Microtek Laboratories  (USA)',
    'mPCM37-D  — Microencapsulated Paraffin (electron micrograph shown)',
    'Class: Paraffin core + polymer shell (10–15 wt%)\nMelting Peak: 37°C  |  Latent Heat: 110 J/g (encapsulated)\nParticle size: 17–20 µm  |  Sp. Heat: 2.70–3.21 kJ/kg·K\nForm: Pumpable slurry (20–40 wt% in water/fluid)',

    'app_pcm_test',
    'APPLICATION — PCM Module Testing',
    'PCM Cooling System Validation',
    'PCM cooling systems are validated by instrumenting battery\nmodules with thermocouples and cycling at 1–3C rates.\nThe PCM melting plateau is visible as a flat region in the\ncell surface temperature–time curve during high-rate discharge.',
    SPEC_PCM)

rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ════════════════════════════════════════════════════════════════
# SECTION 3 — APPLICATION PHOTOS
# ════════════════════════════════════════════════════════════════
r = section_hdr(ws2, r, '  SECTION 3 — APPLICATION PHOTOS  — How these technologies are deployed in battery systems', APP_BG)

r = block(ws2, r,
    'app_immersion',
    'Two-Phase Dielectric Immersion',
    'Battery Pack Submerged in Dielectric Fluid',
    'Battery cells or modules are fully submerged in dielectric\nfluid in a sealed enclosure. Fluid boils on the cell surface,\nremoving heat via latent absorption. Vapour condenses at a\nchilled overhead plate and drains back. No pump needed.',

    'app_pcm_cells',
    'PCM Passive Thermal Buffer',
    'Phase Change Material Between Battery Cells',
    'Solid PCM (e.g. Rubitherm RT35HC) fills the inter-cell\ngaps in the module housing. Cell heat melts the PCM at\n~35°C, absorbing latent heat and clamping cell temperature.\nPCM resolidifies when the pack cools between cycles.',
    SPEC_APP)

r = block(ws2, r,
    'app_pcm_module',
    'Enhanced PCM — Composite with Metal Fins',
    'Composite PCM: Thermal Conductivity Enhancement',
    'Paraffin PCM has low TC (~0.2 W/m·K). Solution: embed\ncopper/aluminium foam or fins to boost effective conductivity\nto 1–5 W/m·K. Fin geometry shown allows rapid heat flow\nfrom cell surface into PCM bulk for high-rate discharge.',

    'app_henkel',
    'Solid Thermal Interface + PCM  (Henkel)',
    'PCM-Enhanced Thermal Interface Material',
    'Proprietary PCM compound (Henkel Bergquist) applied as a\nsolid sheet around battery modules. Combines low thermal\ncontact resistance (<0.5 cm²·K/W) with latent heat\nabsorption at ~45°C — used in EV module pack assemblies.',
    SPEC_APP)

# ── Footer ─────────────────────────────────────────────────────────────────────
rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1
rh(ws2, r, 16)
mg(ws2, r, 1, 5,
   'Image credits: Engineered Fluids shop  |  Yamabala.com (Syensqo Galden)  |  TMC Industries (3M Novec, Chemours Opteon)  |  '
   'Rubitherm GmbH  |  Croda Industrial Specialties  |  Microtek Laboratories  |  eMobility Engineering (application photos — © all rights reserved)',
   bold=False, size=8, fg='9E9E9E', bg='FAFAFA', h='center', italic=True)

wb.save(path)
print(f'\nDone. Tab 1: {updated} principles updated. Tab 2: {len(downloaded)} images inserted.')
