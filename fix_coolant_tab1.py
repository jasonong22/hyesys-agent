import sys, os, urllib.request, tempfile
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = (r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD'
        r'\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\coolants'
        r'\Phase Change Coolant Comparison.xlsx')

# ══════════════════════════════════════════════════════════════════════════════
# CONCISE WORKING PRINCIPLES — temperature/boiling/melting focus, no formulas
# ══════════════════════════════════════════════════════════════════════════════
PRINCIPLES = {
    'Novec 7000':
        'Boils at 34°C — at room temperature, ideal for the 20–40°C battery operating window. '
        'Liquid submerges cells; cell heat causes nucleate boiling on the surface. '
        'Vapour rises to an overhead condenser, condenses, and recirculates by gravity — no pump required.',

    'Novec 7100':
        'Boils at 61°C, clamping peak cell surface temperature near 61°C during fast charging. '
        'Condenser can use standard chilled-water supply (40–55°C), making it directly compatible '
        'with data-centre HVAC. Most widely deployed Novec grade for Li-ion battery immersion cooling.',

    'Novec 7200':
        'Boils at 76°C — suited for high-ambient environments (≥40°C) or cells reaching 60–75°C. '
        'Provides 10–15°C headroom above normal peak cell temperature, preventing parasitic boiling '
        'during standard discharge cycles. Best GWP in the Novec two-phase line (GWP 59).',

    'Novec 7300':
        'Boils at 98°C — above Li-ion safe cell temperature limits; not for direct cell immersion. '
        'Deployed for BMS power electronics (inverters, DC-DC converters) operating at 85–95°C '
        'junction temperatures. Same pool-boiling mechanism as lower-boiling HFE grades.',

    'Novec 649':
        'Boils at 49°C, absorbing heat as latent energy at the cell surface. '
        'At sufficient vapour concentration in an enclosure, simultaneously acts as a fire suppressant. '
        'Being discontinued by Solventum; Opteon 2P50 is the designated commercial successor.',

    'FC-72':
        'Boils at 56°C via nucleate pool boiling on cell surfaces. '
        'Inert to all materials; heat removed entirely as latent energy at the boiling surface. '
        'GWP ~9,300 — prohibited for new installations under current F-gas regulations; '
        'listed for legacy reference only.',

    'FC-3283':
        'Boils at 128°C — above Li-ion thermal runaway onset (~150°C); not suitable for battery cell cooling. '
        'Used for power electronics thermal testing and semiconductor burn-in chambers requiring '
        'fluid stability above 100°C. GWP >10,000; legacy reference only.',

    'Galden EV55':
        'Boils at 55°C, closely matching the Li-ion thermal operating window. '
        'Stays chemically inert to all battery materials including electrolyte solvents and binder '
        'systems. Designed specifically for EV battery direct-immersion cooling.',

    'Galden EV110':
        'Boils at 110°C. During normal operation, stays liquid and removes heat by convection. '
        'If cells approach thermal runaway and surface temperature exceeds 110°C, boiling engages — '
        'absorbing additional latent heat as a secondary containment layer that slows runaway propagation.',

    'Galden HT135':
        'Boils at 135°C — targets BMS power electronics at 100–130°C junction temperatures, '
        'not battery cells directly. '
        'Heat removed as latent energy at the electronics surface; same pool-boiling mechanism as '
        'lower-boiling PFPE grades.',

    'Opteon SF33':
        'Boils at 33°C — almost identical to Novec 7000 (34°C), enabling a drop-in replacement '
        'with no hardware modifications. '
        'Designed as the low-GWP successor (GWP 2) as Novec 7000 is discontinued by Solventum. '
        'Vapour condenses at overhead cooler and recirculates by gravity.',

    'Opteon 2P50':
        'Boils at ~49°C, matching the cell surface temperature during moderate charge rates. '
        'GWP 10 — next-generation successor to Novec 649 and PFPE fluids at far lower environmental cost. '
        'First commercial deployments 2025 (NTT DATA, Hibiya Engineering, Samsung-approved electronics).',

    'EC-100':
        'Single-phase; no boiling within the operating range (flash point 190°C). '
        'Heat is removed entirely by forced convection in a pumped recirculation loop. '
        'Higher thermal conductivity (~0.138 W/m·K) than fluorinated fluids compensates for '
        'the absence of latent heat. Zero GWP.',

    'EC-110':
        'Single-phase; no boiling within the operating range (flash point 193°C). '
        'Same forced-convection pumped-loop mechanism as EC-100 but with higher dielectric '
        'strength (≥60 kV) for 800V+ EV platforms. Pour point −57°C enables cold-climate deployment.',

    'AC-110':
        'Single-phase; heat removed by forced convection in a pumped loop. '
        'Formulated for simultaneous cooling of battery cells, motor windings, and inverter '
        'electronics in one fluid circuit, eliminating separate coolant loops and heat exchangers.',

    'RT28HC':
        'Melts at 27–29°C, absorbing 250 J/g of latent heat during phase transition. '
        'Cell surface temperature is held near 28°C throughout melting — preventing heat build-up '
        'at the critical 30–35°C degradation threshold. Resolidifies between cycles passively.',

    'RT35HC':
        'Melts at 34–36°C, absorbing 240 J/g. '
        'Clamps cell surface temperature near 35°C during fast-charging heat bursts, '
        'preventing the accelerated degradation that occurs above 40°C in NMC cells. '
        'Most validated PCM temperature window in battery BTM research literature.',

    'RT44HC':
        'Melts at 41–45°C, absorbing 250 J/g. '
        'Suited for LFP cells or tropical deployments where a 40–45°C operating window '
        'is acceptable before the PCM engages. Equal latent heat to RT28HC at a higher temperature.',

    'RT55':
        'Melts at 51–57°C, absorbing 170 J/g. '
        'Activates in the thermal runaway early-warning zone (>50°C), acting as a passive '
        'heat-absorption buffer that slows temperature rise and thermal propagation to adjacent cells.',

    'RT60':
        'Melts progressively across a wide 53–68°C band, absorbing ~160–180 J/g. '
        'The wide range accommodates large temperature gradients within a pack — '
        'hotter cells engage the PCM earlier while cooler cells have not yet reached the transition.',

    'RT64HC':
        'Melts at 60–65°C, estimated 240–250 J/g (HC high-capacity formulation). '
        'Positioned as a passive heat-absorption buffer near the thermal runaway boundary '
        'for power electronics or as a safety layer in high-temperature BESS modules.',

    'CrodaTherm 37':
        'Melts at 36°C, absorbing 216 J/g — identical working principle to paraffin PCMs. '
        'Cell surface temperature is clamped at 36°C during phase transition. '
        '100% plant-derived (bio-based) renewable carbon; lower flammability than petroleum paraffin.',

    'CrodaTherm 47':
        'Melts at 47°C, absorbing 197 J/g. Passive solid-liquid thermal buffering at 47°C. '
        'Available in microencapsulated (ME) form as a pumpable slurry, enabling integration '
        'into existing liquid-cooled battery systems without module redesign.',

    'CrodaTherm 53':
        'Melts at 52°C, absorbing 220 J/g. Passive thermal buffer for LFP BESS modules '
        'approaching their upper operating limit (~55°C). '
        'ME microencapsulated slurry variants available for liquid-cooled circuit integration.',

    'CrodaTherm 60':
        'Melts at 60°C, absorbing 215 J/g — at the critical boundary where cells approach '
        'thermal runaway early-warning limits. Absorbs heat passively, buying time for BMS '
        'protection systems to respond. ME slurry grades available for active cooling circuits.',

    'MPCM37-D':
        'Phase change peak at 37°C; onset at 25.5°C. '
        'Microcapsules (17–20 µm) are dispersed in a pumped slurry and remove heat both by '
        'fluid convection (flow) and by latent absorption at 37°C. '
        'Heat is released at a remote heat exchanger, not at the cell surface.',
}

# ══════════════════════════════════════════════════════════════════════════════
# DIFFERENTIATED COOLING MECHANISMS
# ══════════════════════════════════════════════════════════════════════════════
MECHANISMS = {
    'Novec 7000':    'Two-Phase / HFE',
    'Novec 7100':    'Two-Phase / HFE',
    'Novec 7200':    'Two-Phase / HFE',
    'Novec 7300':    'Two-Phase / HFE',
    'Novec 649':     'Two-Phase / Fluoroketone',
    'FC-72':         'Two-Phase / PFC (legacy)',
    'FC-3283':       'Two-Phase / PFC (legacy)',
    'Galden EV55':   'Two-Phase / PFPE',
    'Galden EV110':  'Two-Phase / PFPE',
    'Galden HT135':  'Two-Phase / PFPE',
    'Opteon SF33':   'Two-Phase / HFO (low-GWP)',
    'Opteon 2P50':   'Two-Phase / HFO (low-GWP)',
    'EC-100':        'Single-Phase / Synthetic HC',
    'EC-110':        'Single-Phase / Synthetic HC',
    'AC-110':        'Single-Phase / Synthetic HC',
    'RT28HC':        'Passive PCM / Paraffin',
    'RT35HC':        'Passive PCM / Paraffin',
    'RT44HC':        'Passive PCM / Paraffin',
    'RT55':          'Passive PCM / Paraffin',
    'RT60':          'Passive PCM / Paraffin',
    'RT64HC':        'Passive PCM / Paraffin',
    'CrodaTherm 37': 'Passive PCM / Bio-fatty acid',
    'CrodaTherm 47': 'Passive PCM / Bio-fatty acid',
    'CrodaTherm 53': 'Passive PCM / Bio-fatty acid',
    'CrodaTherm 60': 'Passive PCM / Bio-fatty acid',
    'MPCM37-D':      'PCM Slurry / Recirculating',
}

# ══════════════════════════════════════════════════════════════════════════════
# LOAD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = load_workbook(PATH)
ws1 = wb['Coolant Comparison']

# Find header row by scanning for 'Vendor' in col 2
header_row = None
for ri in range(1, 10):
    if ws1.cell(row=ri, column=2).value == 'Vendor':
        header_row = ri
        break
if header_row is None:
    raise RuntimeError('Could not find header row with "Vendor" in column 2')
print(f'Header row: {header_row}')

# Get column indices from header row
headers_raw = [ws1.cell(row=header_row, column=c).value for c in range(1, 18)]
def col_of(name):
    return headers_raw.index(name) + 1  # 1-based

vendor_col = col_of('Vendor')
prod_col   = col_of('Product / Grade')
mech_col   = col_of('Cooling Mechanism')
princ_col  = col_of('Working Principle')
print(f'Columns — Vendor:{vendor_col} Product:{prod_col} Mechanism:{mech_col} Principle:{princ_col}')

# ══════════════════════════════════════════════════════════════════════════════
# PASS 1 — update cells, collect data rows for vendor merging
# ══════════════════════════════════════════════════════════════════════════════
data_rows = []   # list of (row_idx, vendor_string)
updated_p = 0
updated_m = 0

for ri in range(header_row + 1, ws1.max_row + 1):
    vendor_cell = ws1.cell(row=ri, column=vendor_col)
    prod_cell   = ws1.cell(row=ri, column=prod_col)

    vval = vendor_cell.value
    pval = prod_cell.value
    if not vval or not pval:
        continue  # blank or section-header row

    vstr = str(vval)
    pstr = str(pval)

    # Fix vendor name — remove "(Solvay)" from Syensqo
    if 'Solvay' in vstr:
        vendor_cell.value = 'Syensqo'
        vstr = 'Syensqo'

    # Update cooling mechanism
    for key, mech_val in MECHANISMS.items():
        if key in pstr:
            ws1.cell(row=ri, column=mech_col).value = mech_val
            updated_m += 1
            break

    # Update working principle
    for key, principle in PRINCIPLES.items():
        if key in pstr:
            ws1.cell(row=ri, column=princ_col).value = principle
            updated_p += 1
            break

    data_rows.append((ri, vstr))

print(f'Updated: {updated_p} principles, {updated_m} mechanisms, {len(data_rows)} data rows')

# ══════════════════════════════════════════════════════════════════════════════
# PASS 2 — merge consecutive same-vendor cells vertically
# ══════════════════════════════════════════════════════════════════════════════
# First, unmerge any existing merges in the vendor column for safety
for merge_range in list(ws1.merged_cells.ranges):
    if merge_range.min_col == vendor_col and merge_range.max_col == vendor_col:
        ws1.unmerge_cells(str(merge_range))

# Group consecutive same-vendor rows
from itertools import groupby

def grouper(rows):
    for vendor, group in groupby(rows, key=lambda x: x[1]):
        group_list = list(group)
        yield vendor, [r for r, _ in group_list]

merged_count = 0
for vendor, row_indices in grouper(data_rows):
    if len(row_indices) < 2:
        continue
    r_start = row_indices[0]
    r_end   = row_indices[-1]
    top_cell = ws1.cell(row=r_start, column=vendor_col)
    top_val  = top_cell.value   # capture value before merge resets it
    ws1.merge_cells(start_row=r_start, start_column=vendor_col,
                    end_row=r_end,   end_column=vendor_col)
    # Re-apply value and vertical-centre alignment to merged top cell
    merged_top = ws1.cell(row=r_start, column=vendor_col)
    merged_top.value     = top_val
    merged_top.alignment = Alignment(horizontal='center', vertical='center',
                                     wrap_text=True)
    merged_count += 1
    print(f'  Merged vendor "{vendor}" rows {r_start}–{r_end}')

print(f'Vendor merge: {merged_count} groups merged')

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — REBUILD PRODUCT IMAGES TAB
# ══════════════════════════════════════════════════════════════════════════════
def fill_s(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, size=10, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def aln(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin_s = Side(style='thin', color='CFD8DC')
def bdr(): return Border(top=thin_s, bottom=thin_s, left=thin_s, right=thin_s)

def ce(ws, row, col, val='', bold=False, size=10, fg='000000', bg=None,
       italic=False, h='center', v='center', wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap)
    if bg: c.fill = fill_s(bg)
    c.border = bdr()
    return c

def mg(ws, r, c1, c2, val='', bold=False, size=10, fg='000000', bg=None,
       h='center', italic=False, v='center'):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=val)
    c.font      = fnt(bold, size, fg, italic)
    c.alignment = aln(h, v, wrap=True)
    if bg: c.fill = fill_s(bg)
    c.border = bdr()
    return c

def rh(ws, row, h): ws.row_dimensions[row].height = h
def bg_row(ws, row, bg, ncol=5):
    for col in range(1, ncol+1):
        ws.cell(row=row, column=col).fill = fill_s(bg)

if 'Product Images' in wb.sheetnames:
    del wb['Product Images']
ws2 = wb.create_sheet('Product Images', 1)
ws2.sheet_view.showGridLines = False

col_w = {1:3, 2:44, 3:3, 4:44, 5:3}
for col, w in col_w.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

# Download images
TMPDIR = tempfile.mkdtemp(prefix='coolant_imgs_')
downloaded = {}
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

def dl(key, url, ext='.jpg'):
    fpath = os.path.join(TMPDIR, f'{key}{ext}')
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15) as r, open(fpath, 'wb') as f:
            f.write(r.read())
        if os.path.getsize(fpath) > 2000:
            downloaded[key] = fpath
            print(f'  OK  {key}: {os.path.getsize(fpath)//1024} kB')
        else:
            print(f'  SKIP {key}: too small')
    except Exception as e:
        print(f'  FAIL {key}: {e}')

print('\nDownloading images...')
dl('novec_7100',     'https://tmcindustries.com/cdn/shop/files/240626-3M_Novec_7100-Glass-Gallon.jpg?v=1719433544')
dl('galden_ht55',    'https://www.yamabala.com/upload/catalog_list_pic/enL_catalog_25F12_K6BY7iZv3z.jpg')
dl('opteon_sf33',    'https://tmcindustries.com/cdn/shop/files/SF33.png?v=1723479929&width=750', '.png')
# Opteon 2P50 — no product photo publicly available yet; reuse SF33 (same product line)
dl('opteon_2p50',    'https://tmcindustries.com/cdn/shop/files/SF33.png?v=1723479929&width=750', '.png')
dl('ampcool',        'https://shop.engineeredfluids.com/cdn/shop/products/AC-110B_web.jpg?v=1675264551')
dl('electrocool',    'https://shop.engineeredfluids.com/cdn/shop/products/EC-110_web.jpg?v=1675264540')
dl('rubitherm_rt',   'https://www.rubitherm.eu/media/products/images/_detailImage/Rubitherm-RT-400x240.jpg')
dl('croda',          'https://www.crodaindustrialspecialties.com/mediaassets/images/industrial-chemicals/brochure-front-covers/screenshot-20240215-132834.png?w=500&la=en-GB', '.png')
dl('microtek_micro', 'https://microteklabs.com/assets/microcapsules-micrograph-DjIT-j5b.jpg')
dl('app_immersion',  'https://www.emobility-engineering.com/content/uploads/2024/02/EME-Carrar-min-1024x640.jpg')
dl('app_pcm_cells',  'https://www.emobility-engineering.com/content/uploads/2024/02/EME-paraffin-cells-min-1024x331.jpg')
dl('app_pcm_module', 'https://www.emobility-engineering.com/content/uploads/2024/02/EME-compsite-PCM-min-1024x576.jpg')
dl('app_henkel',     'https://www.emobility-engineering.com/content/uploads/2024/02/EME-Henkel-min-1024x768.jpg')
dl('app_pcm_test',   'https://www.emobility-engineering.com/content/uploads/2024/02/PCM-Image-1024x389.jpg')

TITLE_BG = '0D47A1'
DIEL_BG  = '1565C0'
PCM_BG   = '2E7D32'
APP_BG   = '4A148C'
LABEL_BG = 'ECEFF1'
IMG_BG   = 'F5F5F5'
SPEC_BG  = 'E8EAF6'
SPEC_PCM = 'E8F5E9'
SPEC_APP = 'F3E5F5'
WHITE    = 'FFFFFF'
NOTE_FG  = '546E7A'

IMG_H = 160
IMG_W = 285
IMG_D = 160

def insert_img(ws, key, anchor, w=IMG_W, h=IMG_D):
    if key not in downloaded:
        return
    try:
        img = XLImage(downloaded[key])
        img.width  = w
        img.height = h
        img.anchor = anchor
        ws.add_image(img)
    except Exception as e:
        print(f'  IMG FAIL {key}: {e}')

def section_hdr(ws, r, text, bg):
    rh(ws, r, 28)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = fnt(True, 12, WHITE)
    c.alignment = aln('left', 'center')
    c.fill      = fill_s(bg)
    return r + 1

def block(ws, r, lkey, lven, lprod, lspec, rkey, rven, rprod, rspec, sbg):
    rh(ws, r, IMG_H)
    ws.cell(row=r, column=2).fill = fill_s(IMG_BG)
    ws.cell(row=r, column=4).fill = fill_s(IMG_BG)
    if lkey: insert_img(ws, lkey, f'B{r}')
    if rkey: insert_img(ws, rkey, f'D{r}')
    r += 1
    rh(ws, r, 18)
    ce(ws, r, 2, lven, bold=True, size=9, fg='1A237E', bg=LABEL_BG)
    ce(ws, r, 4, rven, bold=True, size=9, fg='1A237E', bg=LABEL_BG)
    r += 1
    rh(ws, r, 18)
    ce(ws, r, 2, lprod, bold=True, size=10, fg='000000', bg=WHITE)
    ce(ws, r, 4, rprod, bold=True, size=10, fg='000000', bg=WHITE)
    r += 1
    rh(ws, r, 40)
    ce(ws, r, 2, lspec, bold=False, size=9, fg='212121', bg=sbg, v='top')
    ce(ws, r, 4, rspec, bold=False, size=9, fg='212121', bg=sbg, v='top')
    r += 1
    rh(ws, r, 8); bg_row(ws, r, 'FAFAFA'); r += 1
    return r

r = 1
rh(ws2, r, 36)
mg(ws2, r, 1, 5, 'Phase Change Coolants — Product Images & Application Photos',
   bold=True, size=15, fg=WHITE, bg=TITLE_BG)
r += 1
rh(ws2, r, 18)
mg(ws2, r, 1, 5,
   'Product images: vendor & distributor websites  |  '
   'Application photos: eMobility Engineering (© all rights reserved)  |  '
   'HyESys Agent  |  June 2026',
   bold=False, size=8, fg=NOTE_FG, bg='ECEFF1', italic=True)
r += 1
rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ── SECTION 1: DIELECTRIC FLUIDS ─────────────────────────────────────────────
r = section_hdr(ws2, r, '  SECTION 1 — TWO-PHASE IMMERSION DIELECTRIC FLUIDS', DIEL_BG)

r = block(ws2, r,
    'novec_7100',  '3M / Solventum',  'Novec 7100  (HFE-7100)',
    'Class: HFE  |  Boiling Pt: 61°C  |  Latent Heat: 112 kJ/kg\n'
    'TC liq: 0.069 W/m·K  |  Density: 1.51 g/cm³\n'
    'GWP: 297–320  |  Dielectric: >40 kV\n'
    'Note: Novec 7000 (34°C) and 7200 (76°C, GWP 59) also available',

    'galden_ht55',  'Syensqo',  'Galden EV55 / HT55  (PFPE)',
    'Class: PFPE  |  Boiling Pt: 55°C  |  Latent Heat: —\n'
    'TC liq: 0.065 W/m·K  |  Density: 1.65 g/cm³\n'
    'GWP: ~10,000* (not officially disclosed)  |  Dielectric: 40 kV\n'
    'EV55 and HT55 share same bottle format shown here',
    SPEC_BG)

r = block(ws2, r,
    'opteon_sf33',  'Chemours',  'Opteon SF33  (HFO, GWP 2)',
    'Class: HFO  |  Boiling Pt: 33°C  |  GWP: 2\n'
    'Drop-in replacement for Novec 7000  |  ODP: 0  |  Non-flammable\n'
    'Chemours expanding production capacity for 2025+ deployments',

    'opteon_2p50',  'Chemours',  'Opteon 2P50  (HFO blend, GWP 10)',
    'Class: HFO blend  |  Boiling Pt: ~49°C  |  GWP: 10\n'
    'TC liq: 0.073 W/m·K  |  Sp. Heat: 1.09 kJ/kg·K\n'
    'Successor to Novec 649 & PFPE fluids  |  Commercial 2025\n'
    '⚠ No product packaging photo available yet — SF33 image shown (same product line)',
    SPEC_BG)

r = block(ws2, r,
    'ampcool',  'Engineered Fluids',  'AmpCool AC-110',
    'Class: Synthetic Hydrocarbon  |  Single-phase (pumped)\n'
    'TC: ~0.138 W/m·K  |  Density: ~0.80 g/cm³\n'
    'GWP: 0  |  Biodegradable ≥95%  |  Flash Pt: >180°C\n'
    'Single-fluid EV strategy: battery + motor + inverter cooling',

    'electrocool',  'Engineered Fluids',  'ElectroCool EC-110',
    'Class: Synthetic Hydrocarbon  |  Single-phase (pumped)\n'
    'TC: 0.136 W/m·K  |  Density: 0.82 g/cm³\n'
    'GWP: 0  |  Dielectric: ≥60 kV  |  Pour Pt: −57°C\n'
    'Shelf life 25 yr  |  Higher dielectric for 800V+ platforms',
    SPEC_BG)

rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ── SECTION 2: PHASE CHANGE MATERIALS ────────────────────────────────────────
r = section_hdr(ws2, r,
    '  SECTION 2 — PHASE CHANGE MATERIALS (PCMs)  — Passive solid↔liquid heat absorption',
    PCM_BG)

r = block(ws2, r,
    'rubitherm_rt',  'Rubitherm Technologies GmbH  (Germany)',
    'RT Series  (RT28HC / RT35HC / RT44HC)',
    'Class: Paraffin wax (petroleum-based)\n'
    'Melting Range: 27–45°C  |  Latent Heat: 240–250 J/g\n'
    'TC: 0.2 W/m·K  |  Density solid: 0.88 g/cm³  |  GWP: N/A\n'
    'Most cited PCM brand in battery BTM literature',

    'croda',  'Croda Industrial Specialties  (UK)',
    'CrodaTherm Series  (37 / 47 / 53 / 60)',
    'Class: Bio-fatty acid  |  100% renewable carbon\n'
    'Melting Range: 36–60°C  |  Latent Heat: 197–220 J/g\n'
    'TC: ~0.15–0.20 W/m·K (est.)  |  GWP: N/A\n'
    'USDA Certified Biobased; lower flammability than paraffin',
    SPEC_PCM)

r = block(ws2, r,
    'microtek_micro',  'Microtek Laboratories  (USA)',
    'mPCM37-D — Microencapsulated Paraffin (electron micrograph)',
    'Class: Paraffin core + polymer shell (10–15 wt%)\n'
    'Melting Peak: 37°C (onset 25.5°C)  |  Latent Heat: 110 J/g\n'
    'Particle size: 17–20 µm  |  Sp. Heat slurry: 2.70–3.21 kJ/kg·K\n'
    'Form: Pumpable slurry (20–40 wt% in water or dielectric fluid)',

    'app_pcm_test',  'APPLICATION — PCM Module Testing',
    'PCM Cooling System Validation',
    'PCM cooling systems are validated by cycling battery modules at\n'
    '1–3C rates with thermocouple arrays. The PCM melting plateau\n'
    'appears as a flat region in the cell temperature–time curve\n'
    'during high-rate discharge — confirming latent heat absorption.',
    SPEC_PCM)

rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1

# ── SECTION 3: APPLICATION PHOTOS ────────────────────────────────────────────
r = section_hdr(ws2, r,
    '  SECTION 3 — APPLICATION PHOTOS  — How these technologies are deployed in battery systems',
    APP_BG)

r = block(ws2, r,
    'app_immersion',  'Two-Phase Dielectric Immersion',
    'Battery Pack Submerged in Dielectric Fluid',
    'Battery cells or modules are fully submerged in dielectric\n'
    'fluid in a sealed enclosure. Fluid boils on the cell surface,\n'
    'removing heat via latent absorption. Vapour condenses at a\n'
    'chilled overhead plate and drains back. No pump needed.',

    'app_pcm_cells',  'PCM Passive Thermal Buffer',
    'Phase Change Material Between Battery Cells',
    'Solid PCM (e.g. Rubitherm RT35HC) fills the inter-cell\n'
    'gaps in the module housing. Cell heat melts the PCM at\n'
    '~35°C, absorbing latent heat and clamping cell temperature.\n'
    'PCM resolidifies when the pack cools between cycles.',
    SPEC_APP)

r = block(ws2, r,
    'app_pcm_module',  'Enhanced PCM — Composite with Metal Fins',
    'Composite PCM: Thermal Conductivity Enhancement',
    'Paraffin PCM has low TC (~0.2 W/m·K). Solution: embed\n'
    'copper or aluminium foam to boost effective conductivity\n'
    'to 1–5 W/m·K. Fin geometry allows rapid heat flow\n'
    'from cell surface into PCM bulk for high-rate discharge.',

    'app_henkel',  'PCM-Enhanced Thermal Interface Material  (Henkel)',
    'Solid Thermal Interface + PCM',
    'Proprietary PCM compound (Henkel Bergquist) applied as a\n'
    'solid sheet around battery modules. Combines low thermal\n'
    'contact resistance (<0.5 cm²·K/W) with latent heat\n'
    'absorption at ~45°C — used in EV module pack assemblies.',
    SPEC_APP)

rh(ws2, r, 6); bg_row(ws2, r, 'FAFAFA'); r += 1
rh(ws2, r, 16)
mg(ws2, r, 1, 5,
   'Image credits: Engineered Fluids  |  Yamabala.com (Syensqo Galden)  |  '
   'TMC Industries (3M Novec, Chemours Opteon)  |  Rubitherm GmbH  |  '
   'Croda Industrial Specialties  |  Microtek Laboratories  |  '
   'eMobility Engineering (application photos — © all rights reserved)',
   bold=False, size=8, fg='9E9E9E', bg='FAFAFA', h='center', italic=True)

wb.save(PATH)
print(f'\nDone. Tab 1: {updated_p} principles, {updated_m} mechanisms updated, '
      f'{merged_count} vendor groups merged. '
      f'Tab 2: {len(downloaded)} images downloaded and inserted.')
