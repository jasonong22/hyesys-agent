import sys, math, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment as XLComment

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Singapore DC Battery Cell Comparison.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb['Cell Comparison']
headers = [cell.value for cell in ws[1]]

model_col  = headers.index('Model')
energy_col = headers.index('Energy (Wh)')
weight_col = headers.index('Weight (g)')
whkg_col   = headers.index('Wh/kg')

# Yellow highlight for corrected cells
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_font    = Font(color='C00000', bold=True)

corrections = {
    '72174L4-280Ah': 163,
    '72174L4-314Ah': 181,
    'LF168':         174,
}

for row_idx in range(2, ws.max_row + 1):
    model = ws.cell(row=row_idx, column=model_col + 1).value
    if model not in corrections:
        continue

    energy_raw = ws.cell(row=row_idx, column=energy_col + 1).value
    weight_raw = ws.cell(row=row_idx, column=weight_col + 1).value
    old_val    = ws.cell(row=row_idx, column=whkg_col + 1).value
    new_val    = corrections[model]

    cell = ws.cell(row=row_idx, column=whkg_col + 1)
    cell.value = new_val
    cell.fill  = yellow_fill
    cell.font  = red_font

    # Add cell comment documenting the correction
    note_text = (
        f"CORRECTED: was {old_val}, now {new_val}\n"
        f"Energy={energy_raw} Wh / Weight={weight_raw} g\n"
        f"Calc: {energy_raw} / ({weight_raw}/1000) = {new_val} Wh/kg"
    )
    comment = XLComment(note_text, 'HyESys Agent')
    comment.width  = 200
    comment.height = 70
    cell.comment   = comment

    print(f"  {model}: {old_val} -> {new_val}  (highlighted + comment added)")

wb.save(path)
print("Saved.")
