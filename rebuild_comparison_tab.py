import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = r'C:\Users\JasonOng\OneDrive - Advancer Global Ltd\AST BD\2024 HyESys\Hardware (PCS.BATT)\v2.2 - data center\cell\Singapore DC Battery Cell Comparison.xlsx'
wb = openpyxl.load_workbook(path)

if 'Cell Type Comparison' in wb.sheetnames:
    del wb['Cell Type Comparison']
ws = wb.create_sheet('Cell Type Comparison')

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h): return PatternFill('solid', fgColor=h)
def fnt(bold=False, size=11, color='000000', italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name='Calibri')
def aln(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
thin_s = Side(style='thin',   color='CFD8DC')
med_s  = Side(style='medium', color='90A4AE')
def bdr(all_med=False):
    s = med_s if all_med else thin_s
    return Border(top=s, bottom=s, left=s, right=s)

def cell(row, col, val='', bold=False, size=11, color='000000',
         bg=None, italic=False, h='left', v='center', wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = fnt(bold=bold, size=size, color=color, italic=italic)
    c.alignment = aln(h=h, v=v, wrap=wrap)
    if bg: c.fill = fill(bg)
    c.border = bdr()
    return c

def merge(row, c1, c2, val='', bold=False, size=11, color='000000',
          bg=None, h='center', italic=False, v='center'):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row,   end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font      = fnt(bold=bold, size=size, color=color, italic=italic)
    c.alignment = aln(h=h, v=v, wrap=True)
    if bg: c.fill = fill(bg)
    c.border = bdr(all_med=True)
    return c

def rh(row, h): ws.row_dimensions[row].height = h

# ── Colours ───────────────────────────────────────────────────────────────────
TITLE_BG   = '1A237E'   # dark indigo
CYL_HDR    = '1565C0'   # blue
CYL_PROS   = 'BBDEFB'   # light blue
CYL_CONS   = 'E3F2FD'   # very light blue
PRI_HDR    = 'E65100'   # deep orange
PRI_PROS   = 'FFE0B2'   # light orange
PRI_CONS   = 'FFF3E0'   # very light orange
PRO_HDR    = '2E7D32'   # green
CON_HDR    = 'B71C1C'   # red
PRO_TITLE  = 'E8F5E9'   # pale green
CON_TITLE  = 'FFEBEE'   # pale red
PRO_DETAIL = 'F1F8E9'   # near-white green
CON_DETAIL = 'FFF8F8'   # near-white red
NOTE_BG    = 'FFF9C4'   # yellow note
NOTE_HDR   = '33691E'   # dark green
SEP_BG     = 'ECEFF1'   # grey separator column
FOOT_BG    = 'FAFAFA'

# ── Column widths ─────────────────────────────────────────────────────────────
# A=spacer, B=cylindrical, C=divider, D=prismatic, E=spacer
widths = {1: 3, 2: 62, 3: 3, 4: 62, 5: 3}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Separator column style ────────────────────────────────────────────────────
for r in range(1, 80):
    ws.cell(row=r, column=3).fill = fill(SEP_BG)

r = 1  # row counter

# ── TITLE ──────────────────────────────────────────────────────────────────────
rh(r, 38)
merge(r, 1, 5, 'Battery Cell Type Comparison: Cylindrical vs Prismatic',
      bold=True, size=16, color='FFFFFF', bg=TITLE_BG, h='center')
r += 1

rh(r, 18)
merge(r, 1, 5,
      'Side-by-side pros & cons  |  Source: cell specifications in this workbook  |  '
      'Geometry verified from measured dimensions',
      bold=False, size=9, color='546E7A', bg='ECEFF1', h='center', italic=True)
r += 1

# ── COLUMN HEADERS ─────────────────────────────────────────────────────────────
r += 1
rh(r, 30)
cell(r, 2, 'CYLINDRICAL CELLS  (18650 · 21700 · 32140)',
     bold=True, size=13, color='FFFFFF', bg=CYL_HDR, h='center')
cell(r, 4, 'PRISMATIC CELLS',
     bold=True, size=13, color='FFFFFF', bg=PRI_HDR, h='center')
r += 1

rh(r, 18)
cell(r, 2, 'INR18650-2500A  |  INR21700-50E (×2)  |  INR21700-M50LT  |  NCR21700A  |  US21700VTC6A  |  Power Cell 32140 (Na-ion)',
     bold=False, size=8, color='1A237E', bg='E8EAF6', h='center', italic=True)
cell(r, 4, 'NMC 50Ah (CATL)  |  72174L4-280Ah (Batterotech)  |  72174L4-314Ah (Batterotech)  |  LF168 (CATL)',
     bold=False, size=8, color='BF360C', bg='FBE9E7', h='center', italic=True)
r += 1

rh(r, 18)
cell(r, 2, 'Cell-level kWh/m³:  Li-ion 21700: 554–742   |   18650: 554   |   Na-ion 32140: 227',
     bold=False, size=9, color='1565C0', bg='E3F2FD', h='center', italic=True)
cell(r, 4, 'Cell-level kWh/m³:  NMC/LFP prismatic: 345–440',
     bold=False, size=9, color='E65100', bg='FFF3E0', h='center', italic=True)
r += 1

# ─────────────────────────────────────────────────────────────────────────────
# PROS SECTION
# ─────────────────────────────────────────────────────────────────────────────
r += 1
rh(r, 24)
cell(r, 2, '  ✔  PROS', bold=True, size=11, color='FFFFFF', bg=PRO_HDR, h='left')
cell(r, 4, '  ✔  PROS', bold=True, size=11, color='FFFFFF', bg=PRO_HDR, h='left')
r += 1

cyl_pros = [
    ('Highest cell-level volumetric energy density',
     'Li-ion 21700: 554–742 kWh/m³ (cell level) — highest of all cell types in this dataset. '
     'Even after 20–35% packing loss for cooling channels, pack-level density remains above '
     'that of equivalent prismatic packs.'),
    ('Mature high-volume manufacturing',
     'Widest global supplier base. 18650 and 21700 are international standards — '
     'multi-sourcing reduces supply risk and keeps cost per Wh competitive.'),
    ('Strong mechanical casing',
     'Steel can withstands internal pressure from gas generation. Cylindrical geometry '
     'distributes hoop stress uniformly — robust under abuse and thermal runaway events.'),
    ('Radial heat dissipation',
     'Heat generated at cell core radiates outward in all radial directions to the outer '
     'surface — efficient air or liquid cooling around each cell.'),
    ('Established safety data',
     'Extensive UL9540A, IEC 62133, and UN38.3 test data available across multiple '
     'suppliers. Failure modes well-characterised.'),
]

pri_pros = [
    ('High pack-level packing efficiency',
     'Flat surfaces stack with 85–92% volumetric efficiency vs 65–80% for cylindrical '
     'modules. Reduces enclosure and rack footprint for a given pack capacity.'),
    ('Fewer cells per pack — simpler system',
     'A 100 kWh pack using 280Ah LFP cells needs only ~35 cells vs ~5,500 for 21700. '
     'Fewer interconnects, fewer weld joints, and lower BMS channel count.'),
    ('Simpler BMS architecture',
     'Low cell count → straightforward series-parallel topology. Each cell can be '
     'individually monitored with minimal hardware overhead.'),
    ('Industry standard for stationary BESS',
     'Large-format prismatic LFP dominates grid-scale and data centre energy storage '
     'globally (CATL, BYD, EVE, Batterotech). Proven supply chain and installation '
     'procedures.'),
    ('Easier field replacement',
     'Individual cells or modules are large and accessible. A failing cell in a BESS '
     'cabinet can be swapped without disassembling hundreds of interconnected '
     'cylindrical sub-packs.'),
    ('LFP thermal stability (dataset cells)',
     'LFP thermal runaway onset ~270°C vs ~150–180°C for NMC cylindrical cells. '
     'Reduced risk of thermal propagation in a multi-cell rack.'),
]

max_pros = max(len(cyl_pros), len(pri_pros))
for i in range(max_pros):
    # Title row
    rh(r, 18)
    if i < len(cyl_pros):
        cell(r, 2, f'  ✔  {cyl_pros[i][0]}',
             bold=True, size=10, color='1B5E20', bg=PRO_TITLE)
    else:
        cell(r, 2, '', bg='F5F5F5')
    if i < len(pri_pros):
        cell(r, 4, f'  ✔  {pri_pros[i][0]}',
             bold=True, size=10, color='1B5E20', bg=PRO_TITLE)
    else:
        cell(r, 4, '', bg='F5F5F5')
    r += 1
    # Detail row
    rh(r, 52)
    if i < len(cyl_pros):
        cell(r, 2, f'     {cyl_pros[i][1]}',
             bold=False, size=10, color='212121', bg=PRO_DETAIL, v='top')
    else:
        cell(r, 2, '', bg='F5F5F5')
    if i < len(pri_pros):
        cell(r, 4, f'     {pri_pros[i][1]}',
             bold=False, size=10, color='212121', bg=PRO_DETAIL, v='top')
    else:
        cell(r, 4, '', bg='F5F5F5')
    r += 1

# ─────────────────────────────────────────────────────────────────────────────
# CONS SECTION
# ─────────────────────────────────────────────────────────────────────────────
r += 1
rh(r, 24)
cell(r, 2, '  ✘  CONS', bold=True, size=11, color='FFFFFF', bg=CON_HDR, h='left')
cell(r, 4, '  ✘  CONS', bold=True, size=11, color='FFFFFF', bg=CON_HDR, h='left')
r += 1

cyl_cons = [
    ('Packing inefficiency in modules',
     'Round cross-section leaves 20–35% of module volume as air gaps once cooling '
     'channels and structural spacers are included (geometric minimum 9–22%; '
     'practical with thermal management: 20–35%).'),
    ('High cell count per pack',
     'A 100 kWh pack using 21700 cells (~18 Wh each) requires ~5,500 cells. Each '
     'needs individual interconnects, nickel strips, and a BMS channel — assembly '
     'complexity scales with cell count.'),
    ('BMS complexity',
     'Hundreds to thousands of cells require per-cell or per-group voltage '
     'monitoring. More channels, more firmware logic, and more potential '
     'single-point failure nodes.'),
    ('Harder mid-pack fault detection',
     'A single failing cell buried inside a module is difficult to detect thermally '
     'or electrically without individual cell-level sensors throughout the pack.'),
]

pri_cons = [
    ('Lower cell-level volumetric energy density',
     'Prismatic cells in this dataset: 345–440 kWh/m³ vs 554–742 kWh/m³ for Li-ion '
     'cylindrical. Pack-level packing efficiency partially closes the gap but does '
     'not fully offset the cell-level disadvantage.'),
    ('Cell swelling under cycling',
     'Electrochemical expansion during charge/discharge causes prismatic cells to '
     'swell (LFP ~1–3% per cycle). Packs require mechanical compression fixtures, '
     'expansion gaps, and periodic re-torqueing of end plates.'),
    ('Thermal hotspots in large-format cells',
     'Heat generated at the cell core has a longer conduction path to the cooling '
     'surface in thick cells (e.g. 72mm for 280Ah). Cell centre runs hotter than '
     'edges — limits maximum discharge rate.'),
    ('Less robust casing under internal pressure',
     'Aluminium prismatic cases are less mechanically resistant to internal pressure '
     'spikes than cylindrical steel cans. Vent designs must be carefully engineered '
     'to prevent case rupture on thermal runaway.'),
    ('Larger gas volume per cell in thermal runaway',
     'Per UL9540A test data: Batterotech 280Ah LFP cell vents 129.5 L/cell at '
     '131.1°C. A 1 MWh BESS (91 cells) releases ~11,785 L total gas — critical '
     'design input for ventilation and gas management.'),
    ('Reduced multi-sourcing flexibility',
     'Large-format prismatic form factors are less standardised than 18650/21700. '
     'Switching cell suppliers often requires re-validation of the module mechanical '
     'design and compression fixtures.'),
]

max_cons = max(len(cyl_cons), len(pri_cons))
for i in range(max_cons):
    rh(r, 18)
    if i < len(cyl_cons):
        cell(r, 2, f'  ✘  {cyl_cons[i][0]}',
             bold=True, size=10, color='B71C1C', bg=CON_TITLE)
    else:
        cell(r, 2, '', bg='F5F5F5')
    if i < len(pri_cons):
        cell(r, 4, f'  ✘  {pri_cons[i][0]}',
             bold=True, size=10, color='B71C1C', bg=CON_TITLE)
    else:
        cell(r, 4, '', bg='F5F5F5')
    r += 1
    rh(r, 52)
    if i < len(cyl_cons):
        cell(r, 2, f'     {cyl_cons[i][1]}',
             bold=False, size=10, color='212121', bg=CON_DETAIL, v='top')
    else:
        cell(r, 2, '', bg='F5F5F5')
    if i < len(pri_cons):
        cell(r, 4, f'     {pri_cons[i][1]}',
             bold=False, size=10, color='212121', bg=CON_DETAIL, v='top')
    else:
        cell(r, 4, '', bg='F5F5F5')
    r += 1

# ─────────────────────────────────────────────────────────────────────────────
# ENERGY DENSITY CLARIFICATION NOTE
# ─────────────────────────────────────────────────────────────────────────────
r += 1
rh(r, 26)
merge(r, 1, 5, '  KEY CLARIFICATION — ENERGY DENSITY',
      bold=True, size=11, color='FFFFFF', bg=NOTE_HDR, h='left')
r += 1

notes = [
    ('Cell-level kWh/m³  (from this workbook)',
     'Cylindrical 21700 Li-ion: 554–742 kWh/m³     Prismatic NMC/LFP: 345–440 kWh/m³\n'
     'Cylindrical cells have higher cell-level volumetric density — confirmed by '
     'geometry calculations from measured dimensions.'),
    ('Pack-level kWh/m³  (estimated)',
     'Cylindrical at 65–80% packing efficiency: ~360–590 kWh/m³\n'
     'Prismatic at 85–92% packing efficiency:  ~294–405 kWh/m³\n'
     'Cylindrical packs remain comparable or higher in volumetric density even after packing losses.'),
    ('Why prismatic dominates BESS despite lower density',
     'System simplicity: fewer cells, fewer interconnects, simpler BMS, easier '
     'maintenance, and established BESS supply chain — not energy density.'),
]

for label, detail in notes:
    rh(r, 18)
    merge(r, 1, 5, f'  {label}',
          bold=True, size=10, color='1B5E20', bg=NOTE_BG, h='left')
    r += 1
    rh(r, 52)
    merge(r, 1, 5, f'     {detail}',
          bold=False, size=10, color='212121', bg='FAFFD6', h='left', v='top')
    r += 1

# ── FOOTER ─────────────────────────────────────────────────────────────────────
r += 1
rh(r, 16)
merge(r, 1, 5,
      'Source: Cell Comparison tab (this workbook)  |  UL9540A Report (Batterotech 280Ah)  |  Prepared by HyESys Agent',
      bold=False, size=8, color='9E9E9E', bg=FOOT_BG, h='center', italic=True)

# Freeze top 2 rows
ws.freeze_panes = 'B4'

wb.save(path)
print(f'Done. Tab written with {r} rows.')
